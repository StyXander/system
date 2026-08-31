"""案例注册表、证据闸门、标准模板与安全导入边界。

所有公司共用同一导入协议，不允许增加公司专用后端分支。
内置标准股份只是预登记案例，不能成为接口层的唯一允许对象。
新案例必须通过案例清单登记公司、时点、口径和来源文件。
导入器只读取已结构化字段，不从任意年报自由猜测财务数字。
来源文件必须放在案例包规定目录，不能引用本机绝对路径。
压缩包路径在解压前标准化，拒绝盘符、反斜杠和路径穿越。
压缩包文件数量、解压体积和压缩比均设上限，防止资源滥用。
加密压缩包无法审查内容，因此在导入阶段直接拒绝。
案例编号只允许稳定字符，防止编号被解释为目录路径。
案例编号一旦存在不会静默覆盖，避免新资料破坏旧版本链。
每份年报都要登记披露日期、官方链接和整份文件哈希。
导入时重新计算文件哈希，不能只信任案例清单中的声明值。
文件哈希不一致属于来源失败，案例不得部分导入后继续运行。
来源披露日晚于案例时点时，相关字段不能进入当期证据包。
案例时点控制的是可见信息边界，不代表资料专业上已经充分。
金额单位只允许有限枚举，防止不同单位在计算中静默混用。
报表口径必须统一为合并或母公司，不能跨口径拼接增长率。
币种保留在案例上下文，当前规则不会自动执行汇率换算。
公开、授权脱敏与合成案例必须显式区分，不靠文件名猜测。
公开资料仍需个人信息检查，公开属性不等于可以任意传播。
非公开资料必须同时确认授权与脱敏，否则导入流程关闭。
文本文件会扫描高风险个人信息，年报 PDF 不做不可靠的文本误判。
字段证据编号在案例内保持唯一，便于模型和人工共同引用。
财务字段必须绑定已登记文档、页码、定位说明和口径说明。
营业收入主字段使用报表披露值，不允许用模型生成估计金额。
应收账款主字段必须明确账面余额或净额，不能省略计量基础。
R1 正式比较优先使用账面余额，避免坏账准备变动扭曲增速。
坏账准备作为增强证据独立登记，不覆盖应收账款主字段。
应收账款净额作为勾稽证据独立登记，不参与主口径增长比较。
账面余额减坏账准备应与净额一致，差异必须由人工解释。
R1 至少需要连续两个年度的营业收入和应收账款主字段。
三年趋势只在三个连续期间均登记完整时开放，不能补造缺年。
辅助规则缺少现金流或净利润字段时可以跳过，但必须显示边界。
公开账龄以独立案例证据保存，不冒充客户级账龄明细。
公开账龄每个区间都绑定原年报文档和 PDF 页码。
公开账龄合计必须与登记的应收账款账面余额完成程序勾稽。
账龄勾稽通过只说明转录结构一致，不说明客户余额真实存在。
期后回款为空时保留资料缺口，不用合成数据冒充公开事实。
合同结算条款为空时保留资料缺口，不从通用会计政策推断客户条款。
信用政策未取得时必须保持未验证，不因行业相同而自动补齐。
案例包的增强证据进入运行证据包，但不会改写原始报表数字。
增强证据关闭同名资料缺口时仍保留人工专业复核状态。
标准案例模板本身默认禁止模型传输；项目级真实许可只对公开案例生效。
公开案例可以声明允许模型传输，但真实调用仍受运行模式控制。
模型传输许可属于案例级属性，不由前端复选框临时绕过。
案例来源存储按案例编号隔离，跨案例文档引用必须返回未登记。
RAG 索引目录同样按案例编号隔离，避免检索结果相互污染。
来源下载接口只允许案例清单中的文档编号，不能读取任意文件。
旧标准股份来源链接只读兼容，不作为新案例的接入方式。
案例详情接口公开必要元数据，不公开本机真实存储绝对路径。
公开来源记录保留哈希和页码，使评委能够回到原文复核。
字段结构校验通过不等于专业口径签字，状态必须保留待确认。
案例技术导入通过不等于合规负责人已经冻结正式样例。
第二案例的同行业属性来自公开信息，不能据此声称风险可比。
R1 未触发只说明未出现规定增长错配，不代表公司没有其他风险。
导入失败时只清理由本次创建的隔离目录，不触碰其他案例。
清理目标必须先确认位于案例根目录下，防止异常路径扩大影响。
运行命名空间用于隔离自动化测试与人工运行记录。
测试命名空间不能被正式状态接口当成人工复核证据。
模板内容变更时必须保持案例模式向后兼容或明确升级路径。
新增字段种类时必须同步前缀、口径校验、来源选择和测试。
新增结构化表时必须决定它是主字段、增强证据还是资料缺口。
任何自动提取能力上线前都应保留人工页码确认与原文回查入口。
项目级许可文件只叠加公开资料传输授权，不能放宽私有案例的访问边界。
许可文件损坏或状态未确认时保持原案例策略，不从缺失配置推断同意。
授权确认记录保留确认人、日期和政策来源，便于后续撤销或复验。
运行目录配额覆盖案例与内容存储，防止批量下载耗尽服务磁盘。
配额解析失败时回到受控默认值，不能把无效配置解释为无限空间。
内容寻址 PDF 以整份哈希命名，相同原件在多个案例间只保存一份实体。
内容寻址存储发布前再次计算临时文件哈希，磁盘写入错误必须被发现。
并发下载同一原件使用排他硬链接发布，后到线程不能覆盖已有内容。
已有同名内容若哈希不符立即失败，不能假设文件名足以证明身份。
案例目录在完整写入前可能被并发线程看到，读取方需要等待发布完成。
同一案例同一快照的重复登记可以复用，不同快照必须拒绝静默覆盖。
并发等待设有时限，崩溃遗留的半成品目录不能让请求永久阻塞。
案例写入异常只删除当前调用新建的目录，共享内容存储仍按哈希保持不可变。
字段金额必须是有限数值，字符串无穷值和非数字不能进入 JSON 证据。
比例字段仅接受明确百分比单位，不能继承案例金额单位。
亿元等合法金额单位仍需全案例一致，不能逐字段随意切换数量级。
字段页码必须为正并且不超过登记 PDF 页数，避免生成无法回查的定位。
人工更正同样执行有限数值和页码边界校验，人工输入不享有技术绕过。
人工确认只改变复核决定，不修改原自动候选和来源哈希。
候选内容发生变化时旧人工决定不继承，防止批准错误地跨版本延续。
更正记录同时保存原值和更正值，使导出能够区分机器候选与真人处理。
只有收入与应收账款在同一年度均存在，R1 才把该年度视为技术完整。
人工可用年度还要求两个主字段分别确认，单字段批准不能放行比较。
三年准备状态检查连续窗口，不以三个离散年度冒充趋势资料。
行业专用字段完整性按专用规则要求计算，不强迫金融企业提供普通 R1 字段。
期间来源会排除披露日晚于时点的行，候选存在不代表当时可见。
同一字段存在多条来源时应由前置登记消歧，选择器不按金额大小猜测。
来源文档解析为实际路径后再次检查工作区边界，符号链接不能逃逸案例根。
登记页元数据只返回该页关联证据编号，不把其他页面字段错误附加。
模板 PDF 是合成占位资料，不能被页面描述为真实企业年报。
模板压缩包用于说明协议，默认许可和人工状态必须保持失败关闭。
公开案例自动登记仍保留合法样例确认待处理，下载成功不等于合规冻结。
租户和所有者标识只能由已认证服务端注入，浏览器 manifest 不能自行声明归属。
私有存储后端标记影响访问控制，但不会改变财务字段的来源校验要求。
案例注册的成功边界是资料身份和结构可复验，不是字段真实完整或风险成立。
本模块的核心原则是先证明来源可回查，再允许计算和模型使用。
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import time
import uuid
import zipfile
from copy import deepcopy
from datetime import date, datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Any, Iterable

import pymupdf as fitz
from openpyxl import load_workbook

from . import data as standard_data
from .industry_rules import INDUSTRY_RULES_VERSION, get_specialized_spec
from .schemas import AI_GENERATED_CONTENT_NOTICE


PROJECT_AUTHORIZATION_FILE = "PROJECT_AUTHORIZATION.json"


def _apply_project_authorization(workspace_root: Path, case: dict[str, Any]) -> dict[str, Any]:
    """把项目所有者的公开数据许可叠加到案例，并保留来源快照边界。"""

    # 许可文件属于项目配置；测试临时目录或未配置部署继续保持原来的失败关闭状态。
    path = workspace_root / PROJECT_AUTHORIZATION_FILE
    if not path.is_file() or case.get("sample_type") != "public":
        return case
    try:
        authorization = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return case
    policies = authorization.get("policies") or {}
    if authorization.get("status") != "confirmed" or not policies.get("public_source_model_transfer_allowed"):
        return case

    authorized = deepcopy(case)
    confirmation = {
        "confirmed_by": str(authorization.get("confirmed_by") or "项目所有者（用户）"),
        "confirmed_on": str(authorization.get("confirmed_on") or ""),
        "permission_basis": str(authorization.get("permission_basis") or ""),
        "model_provider": str(authorization.get("model_provider") or "已配置模型供应商"),
        "transmission_scope": str(authorization.get("transmission_scope") or "最小必要证据片段"),
        "approval_reference": str(authorization.get("authorization_id") or "PROJECT-OWNER-AUTH"),
    }
    authorized["model_transfer_allowed"] = True
    authorized["model_transfer_confirmation"] = confirmation
    authorized["legal_sample_confirmation_status"] = "project_owner_authorized_public_source"
    authorized["project_owner_authorization"] = {
        "status": "confirmed",
        "authorization_id": confirmation["approval_reference"],
        "confirmed_by": confirmation["confirmed_by"],
        "confirmed_on": confirmation["confirmed_on"],
        "scope": confirmation["transmission_scope"],
    }

    # 证据确认只在案例编号、来源快照以及可选的案例包哈希全部一致时生效。
    for record in authorization.get("evidence_confirmations") or []:
        if record.get("case_id") != authorized.get("case_id"):
            continue
        if record.get("source_snapshot_id") != authorized.get("source_snapshot_id"):
            continue
        expected_package_hash = str(record.get("package_sha256") or "").upper()
        actual_package_hash = str(authorized.get("package_sha256") or "").upper()
        if expected_package_hash and expected_package_hash != actual_package_hash:
            continue
        authorized["evidence_owner_review_status"] = "owner_confirmed"
        authorized["evidence_owner_confirmation"] = deepcopy(record)
        authorized["source_review_status"] = "owner_confirmed_registered_public_evidence"
        break
    return authorized


CASE_SCHEMA_VERSION = "case_manifest_v1"
TEMPLATE_VERSION = "case-template-v1.0"
MAX_ZIP_BYTES = 50 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 150 * 1024 * 1024
MAX_FILES = 64
MAX_COMPRESSION_RATIO = 120
CASE_ID_PATTERN = re.compile(r"^[A-Z0-9][A-Z0-9_-]{2,39}$")
DOCUMENT_ID_PATTERN = re.compile(r"^[A-Z0-9][A-Z0-9_-]{2,63}$")
SHA256_PATTERN = re.compile(r"^[0-9A-Fa-f]{64}$")
ALLOWED_AMOUNT_UNITS = {"元", "千元", "万元", "百万元", "亿元"}
ALLOWED_SAMPLE_TYPES = {"public", "authorized_deidentified", "synthetic"}
ALLOWED_STATEMENT_SCOPES = {"合并", "母公司"}
ALLOWED_FILE_NAMES = {
    "case_manifest.json",
    "financial_fields.csv",
    "financial_fields.xlsx",
    "aging.csv",
    "aging.xlsx",
    "subsequent_receipts.csv",
    "subsequent_receipts.xlsx",
    "contract_terms.csv",
    "contract_terms.json",
    "deidentification_log.csv",
    "deidentification_log.xlsx",
    "README.md",
    "README_填写说明.txt",
}
ALLOWED_FIELD_KINDS = {
    "revenue",
    "accounts_receivable",
    "accounts_receivable_allowance",
    "accounts_receivable_net",
    "operating_cash_flow",
    "net_profit",
    "contract_assets",
    "long_term_receivables",
    "contract_liabilities",
    "loan_balance",
    "interest_income",
    "nonperforming_loan_ratio",
    "provision_coverage_ratio",
    "insurance_revenue",
    "insurance_service_result",
    "claims_expense",
    "insurance_liabilities",
    "commission_income",
    "margin_financing_assets",
    "impairment_provision",
}
FIELD_PREFIX = {
    "revenue": "REV",
    "accounts_receivable": "AR",
    "accounts_receivable_allowance": "AR_ALLOW",
    "accounts_receivable_net": "AR_NET",
    "operating_cash_flow": "CFO",
    "net_profit": "NP",
    "contract_assets": "CA",
    "long_term_receivables": "LTR",
    "contract_liabilities": "CL",
    "loan_balance": "LOAN",
    "interest_income": "INT",
    "nonperforming_loan_ratio": "NPL",
    "provision_coverage_ratio": "PCR",
    "insurance_revenue": "INS_REV",
    "insurance_service_result": "ISR",
    "claims_expense": "CLAIMS",
    "insurance_liabilities": "INS_LIAB",
    "commission_income": "COMMISSION",
    "margin_financing_assets": "MARGIN",
    "impairment_provision": "IMPAIR",
}
RATIO_FIELD_KINDS = {"nonperforming_loan_ratio", "provision_coverage_ratio"}
ALLOWED_RATIO_UNITS = {"%", "％", "百分比"}


def _runtime_base(workspace_root: Path) -> Path:
    namespace = re.sub(r"[^A-Za-z0-9_-]", "", os.getenv("AUDITTRACE_RUNTIME_NAMESPACE", ""))
    base = workspace_root / "backend" / "runtime"
    return base / namespace if namespace else base


def _resolved_path_text(path: Path) -> str:
    """去除 Windows 可变长路径前缀，同时保留原目录大小写供相对路径使用。"""

    text = str(path.resolve())
    if text.startswith("\\\\?\\UNC\\"):
        text = "\\\\" + text[8:]
    elif text.startswith("\\\\?\\"):
        text = text[4:]
    return os.path.normpath(text)


def _canonical_path_text(path: Path) -> str:
    """统一 Windows 长路径前缀和大小写，供安全边界比较而非展示。"""

    return os.path.normcase(_resolved_path_text(path))


def _path_is_within(path: Path, parent: Path) -> bool:
    """按规范化绝对路径判断包含关系，避免 Windows 并发解析前缀误报。"""

    child_text = _canonical_path_text(path)
    parent_text = _canonical_path_text(parent)
    try:
        return os.path.commonpath((child_text, parent_text)) == parent_text
    except ValueError:
        # 不同盘符或不兼容根路径明确不属于同一安全边界。
        return False


def _relative_path_within(path: Path, parent: Path) -> Path:
    """在完成规范化边界校验后生成稳定相对路径，兼容 Windows 长路径前缀。"""

    if not _path_is_within(path, parent):
        raise ValueError("登记文件路径解析后超出工作区边界。")
    return Path(os.path.relpath(_resolved_path_text(path), _resolved_path_text(parent)))


def _safe_runtime_child(workspace_root: Path, *parts: str) -> Path:
    """解析运行目录子路径，并拒绝符号链接或异常根目录造成的越界。"""

    workspace = workspace_root.resolve()
    runtime = _runtime_base(workspace).resolve()
    if not _path_is_within(runtime, workspace):
        raise ValueError("运行目录解析后超出工作区边界。")
    child = runtime.joinpath(*parts).resolve()
    if not _path_is_within(child, runtime):
        raise ValueError("运行目录子路径解析后越界。")
    return child


def _normalized_server_tenant_id(tenant_id: str | None) -> str | None:
    """规范化由认证层注入的租户编号；浏览器 manifest 不能调用此边界。"""

    if tenant_id is None:
        return None
    normalized = str(tenant_id).strip()
    if not normalized:
        raise ValueError("租户作用域不能为空。")
    if len(normalized) > 200 or any(ord(character) < 32 for character in normalized):
        raise ValueError("租户作用域格式无效。")
    return normalized


def _tenant_runtime_key(tenant_id: str) -> str:
    """租户原值不进入文件名；固定长度哈希同时消除路径穿越与目录泄露。"""

    normalized = _normalized_server_tenant_id(tenant_id)
    if normalized is None:  # pragma: no cover - 调用签名保证非空，保留防御式边界。
        raise ValueError("租户作用域不能为空。")
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest().upper()


def _cases_dir(workspace_root: Path, *, tenant_id: str | None = None) -> Path:
    """返回案例目录；公网私有临时副本按认证租户哈希隔离。"""

    normalized_tenant = _normalized_server_tenant_id(tenant_id)
    path = (
        _safe_runtime_child(workspace_root, "cases")
        if normalized_tenant is None
        else _safe_runtime_child(
            workspace_root,
            "tenant_cases",
            "v1",
            _tenant_runtime_key(normalized_tenant),
            "cases",
        )
    )
    path.mkdir(parents=True, exist_ok=True)
    return path


def _case_dir(workspace_root: Path, case_id: str, *, tenant_id: str | None = None) -> Path:
    """解析单个案例目录，并把最终路径限定在所选租户案例根下。"""

    if not CASE_ID_PATTERN.fullmatch(case_id):
        raise ValueError("案例编号只允许 3—40 位大写字母、数字、下划线或连字符。")
    cases_root = _cases_dir(workspace_root, tenant_id=tenant_id).resolve()
    path = (cases_root / case_id).resolve()
    if not _path_is_within(path, cases_root):
        raise ValueError("案例目录解析后超出当前租户作用域。")
    return path


def _pdf_content_store_dir(workspace_root: Path) -> Path:
    """Return the content-addressed PDF store shared by automatic cases."""

    path = _safe_runtime_child(workspace_root, "pdf_store")
    path.mkdir(parents=True, exist_ok=True)
    return path


def _runtime_quota_bytes() -> int:
    try:
        quota_mb = int(os.environ.get("AUDITTRACE_RUNTIME_QUOTA_MB", "5120"))
    except ValueError:
        quota_mb = 5120
    return max(256, min(quota_mb, 102400)) * 1024 * 1024


def _runtime_size_bytes(workspace_root: Path) -> int:
    total = 0
    runtime = _runtime_base(workspace_root)
    if not runtime.exists():
        return 0
    for path in runtime.rglob("*"):
        if path.is_file():
            try:
                total += path.stat().st_size
            except OSError:
                continue
    return total


def _ensure_content_addressed_pdf(workspace_root: Path, sha256: str, content: bytes) -> Path:
    """Write one immutable PDF blob and return its shared path.

    Case directories receive a hard link to this blob, so repeated official
    reports consume one physical copy while keeping case-local paths stable.
    """

    blob = _pdf_content_store_dir(workspace_root) / f"{sha256.upper()}.pdf"
    if blob.is_file():
        if _sha256_bytes(blob.read_bytes()) != sha256.upper():
            raise ValueError("内容寻址 PDF 已存在但 SHA-256 不一致。")
        return blob
    if _runtime_size_bytes(workspace_root) + len(content) > _runtime_quota_bytes():
        raise ValueError(
            "运行目录已达到 PDF 缓存配额，请清理可恢复的临时运行记录或提高 AUDITTRACE_RUNTIME_QUOTA_MB。"
        )
    # 先写唯一临时文件，再以硬链接的“仅当不存在时创建”语义发布。
    # 多个任务同时下载同一官方 PDF 时不会让读线程看到半个 blob，也不会
    # 让后到线程覆盖先到线程已经被案例目录引用的不可变内容。
    temporary = blob.with_name(f".{blob.stem}.{uuid.uuid4().hex}.tmp")
    try:
        temporary.write_bytes(content)
        if _sha256_bytes(temporary.read_bytes()) != sha256.upper():
            raise ValueError("内容寻址 PDF 临时文件写入后哈希不一致。")
        try:
            os.link(temporary, blob)
        except FileExistsError:
            if _sha256_bytes(blob.read_bytes()) != sha256.upper():
                raise ValueError("内容寻址 PDF 并发发布后 SHA-256 不一致。")
    finally:
        temporary.unlink(missing_ok=True)
    return blob


def _wait_for_registered_case(
    case_dir: Path,
    source_snapshot_id: str,
    *,
    timeout_seconds: float = 5.0,
) -> dict[str, Any] | None:
    """等待同一 case_id 的并发发布完成；目录消失时允许当前调用接管。"""

    deadline = time.monotonic() + timeout_seconds
    while case_dir.exists():
        existing_path = case_dir / "case.json"
        if existing_path.is_file():
            try:
                existing = json.loads(existing_path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                existing = None
            if isinstance(existing, dict):
                if existing.get("source_snapshot_id") != source_snapshot_id:
                    raise ValueError("case_id 已存在且来源快照不同；系统不会静默覆盖旧案例。")
                # 返回副本标记本次没有重写旧字段或人工复核历史。
                result = deepcopy(existing)
                result["reused_existing_case"] = True
                return result
        if time.monotonic() >= deadline:
            raise ValueError("同一 case_id 正在登记但尚未完成，请稍后重试。")
        time.sleep(0.02)
    return None


def _sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest().upper()


def _iso_date(value: Any, field_name: str) -> str:
    text = str(value or "").strip()
    try:
        date.fromisoformat(text)
    except ValueError as error:
        raise ValueError(f"{field_name} 必须是 YYYY-MM-DD。") from error
    return text


def _bool(value: Any, field_name: str) -> bool:
    if isinstance(value, bool):
        return value
    raise ValueError(f"{field_name} 必须是 JSON 布尔值。")


def _safe_zip_name(raw_name: str) -> str:
    """ZIP 内路径在解压前标准化；拒绝盘符、绝对路径、反斜杠和 ``..``。"""
    if not raw_name or "\\" in raw_name or re.match(r"^[A-Za-z]:", raw_name):
        raise ValueError("ZIP 包含不安全路径。")
    path = PurePosixPath(raw_name)
    if path.is_absolute() or any(part in {"", ".", ".."} for part in path.parts):
        raise ValueError("ZIP 包含路径穿越或空路径。")
    normalized = path.as_posix()
    if normalized.startswith("documents/"):
        if len(path.parts) != 2 or path.suffix.lower() != ".pdf":
            raise ValueError("documents 目录只允许直接存放 PDF。")
    elif normalized not in ALLOWED_FILE_NAMES:
        raise ValueError(f"ZIP 包含未允许的文件：{normalized}")
    return normalized


def _read_zip(content: bytes) -> dict[str, bytes]:
    if len(content) > MAX_ZIP_BYTES:
        raise ValueError("案例包超过 50MB 限制。")
    try:
        archive = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile as error:
        raise ValueError("上传文件不是有效 ZIP。") from error
    with archive:
        infos = [item for item in archive.infolist() if not item.is_dir()]
        if not infos or len(infos) > MAX_FILES:
            raise ValueError("案例包文件数量必须在 1—64 之间。")
        total = sum(item.file_size for item in infos)
        if total > MAX_UNCOMPRESSED_BYTES:
            raise ValueError("案例包解压后超过 150MB 限制。")
        result: dict[str, bytes] = {}
        for info in infos:
            name = _safe_zip_name(info.filename)
            if info.flag_bits & 0x1:
                raise ValueError("不接受加密 ZIP。")
            if info.file_size > 0 and info.compress_size == 0:
                raise ValueError("ZIP 压缩信息异常。")
            ratio = info.file_size / max(1, info.compress_size)
            if ratio > MAX_COMPRESSION_RATIO:
                raise ValueError("ZIP 压缩比异常，疑似压缩炸弹。")
            if name in result:
                raise ValueError(f"ZIP 中存在重复路径：{name}")
            result[name] = archive.read(info)
        return result


def _scan_high_risk_personal_information(files: dict[str, bytes]) -> list[str]:
    """扫描可解释的人类文本，跳过哈希、URL、金额和各类技术编号。

    直接对整份 manifest 做正则会把 SHA-256 或公告 URL 中偶然出现的
    11/18 位数字误判成手机号或身份证号。这里按 JSON 键和 CSV 列筛选，
    既保留公司名称、备注、合同摘要等文本检查，也避免技术元数据误报。
    """
    patterns = {
        "居民身份证号": re.compile(r"(?<!\d)\d{17}[0-9Xx](?!\d)"),
        "中国大陆手机号": re.compile(r"(?<!\d)1[3-9]\d{9}(?!\d)"),
        "疑似银行卡号": re.compile(r"(?<!\d)\d{16,19}(?!\d)"),
    }
    technical_keys = {
        "sha256",
        "file_sha256",
        "source_url",
        "case_id",
        "document_id",
        "evidence_id",
        "approval_reference",
        "t0",
        "disclosure_date",
        "retention_expires_at",
        "report_year",
        "year",
        "value",
        "pdf_page",
        "print_page",
    }

    def json_segments(value: Any, key: str = "") -> list[str]:
        if key in technical_keys:
            return []
        if isinstance(value, dict):
            return [segment for child_key, child in value.items() for segment in json_segments(child, str(child_key))]
        if isinstance(value, list):
            return [segment for child in value for segment in json_segments(child, key)]
        return [value] if isinstance(value, str) else []

    findings: list[str] = []
    for name, content in files.items():
        suffix = Path(name).suffix.lower()
        if suffix not in {".json", ".csv", ".md"}:
            continue
        text = content.decode("utf-8-sig", errors="ignore")
        if suffix == ".json":
            try:
                segments = json_segments(json.loads(text))
            except json.JSONDecodeError:
                segments = [text]
        elif suffix == ".csv":
            try:
                rows = csv.DictReader(io.StringIO(text))
                segments = [
                    str(value)
                    for row in rows
                    for key, value in row.items()
                    if key not in technical_keys and value not in (None, "")
                ]
            except csv.Error:
                segments = [text]
        else:
            segments = [text]
        searchable = "\n".join(segments)
        for label, pattern in patterns.items():
            if pattern.search(searchable):
                findings.append(f"{name} 检出{label}")
    return findings


def _required_text(data: dict[str, Any], key: str, label: str, max_length: int = 200) -> str:
    value = str(data.get(key) or "").strip()
    if not value:
        raise ValueError(f"case_manifest.json 缺少 {label}。")
    if len(value) > max_length:
        raise ValueError(f"{label} 过长。")
    return value


def _validate_manifest(
    raw: Any,
    files: dict[str, bytes],
    *,
    allow_scoped_standard_case_id: bool = False,
) -> dict[str, Any]:
    """校验案例清单；只有已认证租户的隔离目录可与内置案例同名。"""

    if not isinstance(raw, dict):
        raise ValueError("case_manifest.json 必须是 JSON 对象。")
    if raw.get("schema_version") != CASE_SCHEMA_VERSION:
        raise ValueError(f"schema_version 必须为 {CASE_SCHEMA_VERSION}。")
    case_id = _required_text(raw, "case_id", "case_id", 40).upper()
    if not CASE_ID_PATTERN.fullmatch(case_id):
        raise ValueError("case_id 只允许 3—40 位大写字母、数字、下划线或连字符。")
    if case_id == standard_data.CASE_ID and not allow_scoped_standard_case_id:
        raise ValueError("不能覆盖内置标准股份案例。")
    amount_unit = _required_text(raw, "amount_unit", "金额单位", 10)
    if amount_unit not in ALLOWED_AMOUNT_UNITS:
        raise ValueError("金额单位只允许元、千元、万元、百万元或亿元。")
    statement_scope = _required_text(raw, "statement_scope", "报表口径", 20)
    if statement_scope not in ALLOWED_STATEMENT_SCOPES:
        raise ValueError("报表口径只允许合并或母公司。")
    sample_type = _required_text(raw, "sample_type", "样例类型", 40)
    if sample_type not in ALLOWED_SAMPLE_TYPES:
        raise ValueError("样例类型必须是 public、authorized_deidentified 或 synthetic。")
    retention_expires_at = _iso_date(raw.get("retention_expires_at"), "保存期限")
    if retention_expires_at < date.today().isoformat():
        raise ValueError("案例保存期限已经届满。")
    model_transfer_allowed = _bool(raw.get("model_transfer_allowed"), "model_transfer_allowed")
    transfer_confirmation: dict[str, str] | None = None
    if model_transfer_allowed:
        # 勾选“来源合法”不自动等于允许外部模型传输。开启模型链前必须把
        # 确认人、依据、供应商和最小传输范围固化在案例包中，便于事后回查。
        raw_confirmation = raw.get("model_transfer_confirmation")
        if not isinstance(raw_confirmation, dict):
            raise ValueError("允许模型传输时必须提供 model_transfer_confirmation 对象。")
        confirmed_on = _iso_date(raw_confirmation.get("confirmed_on"), "模型传输确认日期")
        if confirmed_on > date.today().isoformat():
            raise ValueError("模型传输确认日期不得晚于今天。")
        transfer_confirmation = {
            "confirmed_by": _required_text(raw_confirmation, "confirmed_by", "模型传输确认人", 100),
            "confirmed_on": confirmed_on,
            "permission_basis": _required_text(raw_confirmation, "permission_basis", "模型传输许可依据", 300),
            "model_provider": _required_text(raw_confirmation, "model_provider", "模型供应商", 100),
            "transmission_scope": _required_text(raw_confirmation, "transmission_scope", "最小传输范围", 300),
            "approval_reference": _required_text(raw_confirmation, "approval_reference", "许可记录编号", 120),
        }

    documents_raw = raw.get("documents")
    if not isinstance(documents_raw, list) or not documents_raw:
        raise ValueError("manifest 至少登记一份来源 PDF。")
    documents: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    seen_paths: set[str] = set()
    for item in documents_raw:
        if not isinstance(item, dict):
            raise ValueError("documents 每项必须是对象。")
        document_id = _required_text(item, "document_id", "document_id", 64).upper()
        if not DOCUMENT_ID_PATTERN.fullmatch(document_id) or document_id in seen_ids:
            raise ValueError("document_id 非法或重复。")
        source_file = _required_text(item, "source_file", "source_file", 160)
        source_file = _safe_zip_name(source_file)
        if not source_file.startswith("documents/") or source_file in seen_paths:
            raise ValueError("来源文件必须位于 documents/ 且路径不得重复。")
        if source_file not in files:
            raise ValueError(f"登记来源不存在：{source_file}")
        expected_hash = _required_text(item, "sha256", "来源 SHA-256", 64).upper()
        if not SHA256_PATTERN.fullmatch(expected_hash):
            raise ValueError("来源 SHA-256 格式错误。")
        actual_hash = _sha256_bytes(files[source_file])
        if actual_hash != expected_hash:
            raise ValueError(f"{source_file} 实际哈希与 manifest 不一致。")
        report_year = int(item.get("report_year"))
        if not 2000 <= report_year <= 2100:
            raise ValueError("report_year 超出允许范围。")
        documents.append(
            {
                "document_id": document_id,
                "source_file": source_file,
                "document_type": str(item.get("document_type") or "annual_report")[:40],
                "report_year": report_year,
                "disclosure_date": _iso_date(item.get("disclosure_date"), "披露日期"),
                "source_url": str(item.get("source_url") or "")[:500],
                "sha256": actual_hash,
            }
        )
        seen_ids.add(document_id)
        seen_paths.add(source_file)

    return {
        "schema_version": CASE_SCHEMA_VERSION,
        "case_id": case_id,
        "company_name": _required_text(raw, "company_name", "公司名称"),
        "company_alias": str(raw.get("company_alias") or "")[:100],
        "ticker": str(raw.get("ticker") or "")[:30],
        "t0": _iso_date(raw.get("t0"), "T0"),
        "currency": _required_text(raw, "currency", "币种", 10).upper(),
        "amount_unit": amount_unit,
        "statement_scope": statement_scope,
        "sample_type": sample_type,
        "industry": str(raw.get("industry") or "")[:100],
        "industry_name": str(raw.get("industry_name") or "")[:100],
        "specialized_rule": str(raw.get("specialized_rule") or "")[:100],
        "specialized_required_fields": [
            str(item).strip()
            for item in (raw.get("specialized_required_fields") or [])
            if str(item).strip() in ALLOWED_FIELD_KINDS
        ],
        "model_transfer_allowed": model_transfer_allowed,
        "model_transfer_confirmation": transfer_confirmation,
        "retention_expires_at": retention_expires_at,
        "documents": documents,
    }


def _rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    return [dict(row) for row in csv.DictReader(io.StringIO(text))]


def _rows_from_xlsx(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [str(item or "").strip() for item in values[0]]
    return [dict(zip(headers, row)) for row in values[1:] if any(value not in (None, "") for value in row)]


def _validate_financial_rows(
    rows: list[dict[str, Any]], manifest: dict[str, Any]
) -> list[dict[str, Any]]:
    document_ids = {item["document_id"] for item in manifest["documents"]}
    documents = {item["document_id"]: item for item in manifest["documents"]}
    normalized: list[dict[str, Any]] = []
    seen: set[tuple[str, int]] = set()
    for index, row in enumerate(rows, start=2):
        row_case_id = str(row.get("case_id") or manifest["case_id"]).strip().upper()
        if row_case_id != manifest["case_id"]:
            raise ValueError(f"financial_fields 第 {index} 行发生跨案例串包。")
        kind = str(row.get("field_kind") or "").strip()
        if kind not in ALLOWED_FIELD_KINDS:
            raise ValueError(f"financial_fields 第 {index} 行 field_kind 不受支持。")
        try:
            year = int(row.get("year"))
            value = float(str(row.get("value")).replace(",", ""))
            pdf_page = int(row.get("pdf_page"))
        except (TypeError, ValueError) as error:
            raise ValueError(f"financial_fields 第 {index} 行的年度、金额或 PDF 页码无效。") from error
        if not math.isfinite(value):
            raise ValueError(f"financial_fields 第 {index} 行金额必须是有限数值。")
        document_id = str(row.get("document_id") or "").strip().upper()
        if document_id not in document_ids:
            raise ValueError(f"financial_fields 第 {index} 行引用未登记 document_id。")
        if documents[document_id]["disclosure_date"] > manifest["t0"]:
            raise ValueError(f"financial_fields 第 {index} 行来源披露日晚于 T0。")
        key = (kind, year)
        if key in seen:
            raise ValueError(f"financial_fields 重复登记 {kind}/{year}。")
        seen.add(key)
        unit = str(row.get("unit") or manifest["amount_unit"]).strip()
        if kind in RATIO_FIELD_KINDS:
            if unit not in ALLOWED_RATIO_UNITS:
                raise ValueError(f"financial_fields 第 {index} 行比例字段必须明确使用百分比单位。")
        elif unit != manifest["amount_unit"]:
            raise ValueError(f"financial_fields 第 {index} 行金额单位与 manifest 不一致。")
        statement_scope = str(row.get("statement_scope") or manifest["statement_scope"]).strip()
        if statement_scope != manifest["statement_scope"]:
            raise ValueError(f"financial_fields 第 {index} 行报表口径与 manifest 不一致。")
        locator = str(row.get("locator") or "").strip()
        if not locator:
            raise ValueError(f"financial_fields 第 {index} 行缺少来源定位。")
        basis = str(row.get("field_basis") or ("net" if kind == "accounts_receivable" else "reported"))
        if kind == "accounts_receivable" and basis not in {"gross", "net"}:
            raise ValueError("应收账款 field_basis 只允许 gross 或 net。")
        if kind == "accounts_receivable_allowance" and basis not in {"allowance", "reported"}:
            raise ValueError("坏账准备 field_basis 只允许 allowance 或 reported。")
        if kind == "accounts_receivable_net" and basis not in {"net", "reported"}:
            raise ValueError("应收账款净额 field_basis 只允许 net 或 reported。")
        evidence_id = str(row.get("evidence_id") or f"{manifest['case_id']}_{FIELD_PREFIX[kind]}_{year}").upper()
        if not DOCUMENT_ID_PATTERN.fullmatch(evidence_id):
            raise ValueError(f"financial_fields 第 {index} 行 evidence_id 非法。")
        normalized.append(
            {
                "case_id": manifest["case_id"],
                "evidence_id": evidence_id,
                "field_kind": kind,
                "year": year,
                "value": value,
                "unit": unit,
                "currency": manifest["currency"],
                "statement_scope": statement_scope,
                "field_basis": basis,
                "document_id": document_id,
                "pdf_page": pdf_page,
                "print_page": int(row["print_page"]) if str(row.get("print_page") or "").strip() else None,
                "locator": locator[:300],
            }
        )
    if not normalized:
        raise ValueError("financial_fields 没有数据行。")
    year_kinds: dict[int, set[str]] = {}
    for row in normalized:
        year_kinds.setdefault(row["year"], set()).add(row["field_kind"])
    specialized_required = set(manifest.get("specialized_required_fields") or [])
    required_for_continuity = specialized_required or {"revenue", "accounts_receivable"}
    complete_years = sorted(
        (year for year, kinds in year_kinds.items() if required_for_continuity.issubset(kinds)),
        reverse=True,
    )
    if len(complete_years) < 2:
        rule_label = "行业专用规则" if specialized_required else "R1"
        raise ValueError(f"{rule_label} 至少需要连续两年的完整字段。")
    if any(a - b != 1 for a, b in zip(complete_years, complete_years[1:])):
        rule_label = "行业专用规则" if specialized_required else "R1"
        raise ValueError(f"{rule_label} 基本计算要求至少两年连续期间。")
    return normalized


def _validate_public_case_evidence(
    files: dict[str, bytes],
    manifest: dict[str, Any],
    financial_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str]]:
    """读取案例包中的公开账龄；它只补充解释证据，不覆盖 R1 主字段。"""
    evidence: list[dict[str, Any]] = []
    material_gaps: list[str] = []
    documents = {item["document_id"]: item for item in manifest["documents"]}
    if "aging.csv" in files:
        rows = _rows_from_csv(files["aging.csv"])
        totals: dict[int, float] = {}
        for index, row in enumerate(rows, start=2):
            case_id = str(row.get("case_id") or "").strip().upper()
            if case_id != manifest["case_id"]:
                raise ValueError(f"aging.csv 第 {index} 行发生跨案例串包。")
            try:
                year = int(row.get("year"))
                amount = float(str(row.get("gross_amount")).replace(",", ""))
                pdf_page = int(row.get("pdf_page"))
            except (TypeError, ValueError) as error:
                raise ValueError(f"aging.csv 第 {index} 行的年度、金额或页码无效。") from error
            document_id = str(row.get("document_id") or "").strip().upper()
            if document_id not in documents or documents[document_id]["disclosure_date"] > manifest["t0"]:
                raise ValueError(f"aging.csv 第 {index} 行引用未登记或晚于 T0 的来源。")
            if str(row.get("unit") or "").strip() != manifest["amount_unit"] or amount < 0 or pdf_page < 1:
                raise ValueError(f"aging.csv 第 {index} 行的单位、金额或页码不符合案例口径。")
            bucket = str(row.get("aging_bucket") or "").strip()[:40]
            if not bucket:
                raise ValueError(f"aging.csv 第 {index} 行缺少账龄区间。")
            totals[year] = totals.get(year, 0.0) + amount
            evidence.append(
                {
                    "evidence_id": f"{manifest['case_id']}-AGING-{year}-{index - 1:02d}",
                    "evidence_kind": "public_aging",
                    "field_label": f"{year}年应收账款账龄：{bucket}",
                    "details": {"year": year, "aging_bucket": bucket, "gross_amount": amount, "unit": manifest["amount_unit"]},
                    "document_id": document_id,
                    "pdf_page": pdf_page,
                    "source_url": documents[document_id]["source_url"],
                    "file_sha256": documents[document_id]["sha256"],
                    "source_mode": "case_package_public",
                    "support_status": "import_validated_pending_human_professional_confirmation",
                }
            )
        gross_by_year = {
            row["year"]: row["value"]
            for row in financial_rows
            if row["field_kind"] == "accounts_receivable" and row["field_basis"] == "gross"
        }
        for year, total in totals.items():
            expected = gross_by_year.get(year)
            if expected is None or abs(total - expected) > 0.02:
                raise ValueError(f"aging.csv {year} 年合计与应收账款账面余额不勾稽。")
    if not evidence:
        material_gaps.append("账龄结构")
    if not _rows_from_csv(files.get("subsequent_receipts.csv", b"")):
        material_gaps.append("期后回款")
    if not _rows_from_csv(files.get("contract_terms.csv", b"")):
        material_gaps.extend(["信用政策变动", "主要客户合同结算条款"])
    return evidence, material_gaps


def _standard_case(workspace_root: Path) -> dict[str, Any]:
    documents: dict[str, dict[str, Any]] = {}
    for evidence in standard_data.EVIDENCE.values():
        document_id = f"STD-AR-{evidence['year']}-{evidence['file_sha256'][:12]}"
        documents.setdefault(
            document_id,
            {
                "document_id": document_id,
                "source_file": evidence["source_file"],
                "storage_relpath": evidence["source_file"],
                "document_type": "annual_report",
                "report_year": evidence["year"],
                "disclosure_date": evidence["disclosure_date"],
                "announcement_title": evidence["announcement_title"],
                "source_url": evidence["source_url"],
                "sha256": evidence["file_sha256"],
            },
        )
    case = {
        "schema_version": CASE_SCHEMA_VERSION,
        "case_id": standard_data.CASE_ID,
        "company_name": standard_data.CASE_NAME,
        "company_alias": "标准股份",
        "ticker": standard_data.TICKER,
        "t0": max(item["t0"] for item in standard_data.PERIODS.values()),
        "currency": "CNY",
        "amount_unit": "元",
        "statement_scope": "合并",
        "sample_type": "public",
        # 内置清单保持失败关闭默认值；项目根目录的实名许可记录可在读取时解锁。
        "model_transfer_allowed": False,
        "retention_expires_at": "2026-12-31",
        "legal_sample_confirmation_status": "pending_human_confirmation",
        "source_snapshot_id": standard_data.SOURCE_SNAPSHOT_ID,
        "source_review_status": standard_data.SOURCE_REVIEW_STATUS,
        "documents": list(documents.values()),
        "available_years": sorted(standard_data.PERIODS, reverse=True),
        "three_year_r1_ready": True,
        "registry_mode": "built_in",
    }
    return _apply_project_authorization(workspace_root, case)


def get_case(
    workspace_root: Path,
    case_id: str,
    *,
    tenant_id: str | None = None,
) -> dict[str, Any] | None:
    """读取本地案例；传入租户时只读取该租户的临时物化副本。"""

    if not CASE_ID_PATTERN.fullmatch(case_id):
        return None
    normalized_tenant = _normalized_server_tenant_id(tenant_id)
    # 显式 tenant 是精确私有查询，不能被同 ID 的内置/全局公开案例遮蔽。
    # 无 tenant 才保持本地竞赛模式的内置标准案例兼容合同。
    if case_id == standard_data.CASE_ID and normalized_tenant is None:
        return _standard_case(workspace_root)
    path = _case_dir(workspace_root, case_id, tenant_id=normalized_tenant) / "case.json"
    if not path.is_file():
        return None
    try:
        case = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    # 即使目录内容被本机误拷贝，也不允许把其他租户的清单当作当前副本读取。
    if normalized_tenant is not None and str(case.get("tenant_id") or "").strip() != normalized_tenant:
        return None
    return _apply_project_authorization(workspace_root, case)


def list_cases(workspace_root: Path, *, tenant_id: str | None = None) -> list[dict[str, Any]]:
    """列出本地案例；显式租户查询不扫描其他租户或旧全局私有目录。"""

    normalized_tenant = _normalized_server_tenant_id(tenant_id)
    cases = [] if normalized_tenant is not None else [_standard_case(workspace_root)]
    for path in sorted(_cases_dir(workspace_root, tenant_id=normalized_tenant).glob("*/case.json")):
        try:
            case = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if normalized_tenant is not None and str(case.get("tenant_id") or "").strip() != normalized_tenant:
            continue
        cases.append(_apply_project_authorization(workspace_root, case))
    return cases


def import_case_zip(
    workspace_root: Path,
    content: bytes,
    *,
    authorized: bool,
    desensitized: bool,
    tenant_id: str | None = None,
    owner_user_id: str | None = None,
) -> dict[str, Any]:
    normalized_tenant = _normalized_server_tenant_id(tenant_id)
    if not authorized:
        raise ValueError("必须确认案例资料已获授权或来自合法公开来源。")
    if not desensitized:
        raise ValueError("必须确认非公开资料已脱敏；公开资料也需完成个人信息检查。")
    files = _read_zip(content)
    if "case_manifest.json" not in files:
        raise ValueError("案例包缺少 case_manifest.json。")
    financial_name = "financial_fields.csv" if "financial_fields.csv" in files else "financial_fields.xlsx" if "financial_fields.xlsx" in files else None
    if financial_name is None:
        raise ValueError("案例包缺少 financial_fields.csv 或 financial_fields.xlsx。")
    findings = _scan_high_risk_personal_information(files)
    if findings:
        raise ValueError("案例包检出高风险个人信息：" + "；".join(findings))
    try:
        raw_manifest = json.loads(files["case_manifest.json"].decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as error:
        raise ValueError("case_manifest.json 无法解析。") from error
    manifest = _validate_manifest(
        raw_manifest,
        files,
        allow_scoped_standard_case_id=normalized_tenant is not None,
    )
    raw_rows = _rows_from_csv(files[financial_name]) if financial_name.endswith(".csv") else _rows_from_xlsx(files[financial_name])
    financial_rows = _validate_financial_rows(raw_rows, manifest)
    structured_evidence, material_gaps = _validate_public_case_evidence(files, manifest, financial_rows)
    # tenant_id 只能来自已认证服务端参数；manifest 即使伪造同名字段也已被
    # 白名单解析丢弃。公网副本进入租户哈希目录，普通本地读取不会看到它。
    case_dir = _case_dir(workspace_root, manifest["case_id"], tenant_id=normalized_tenant)
    if case_dir.exists():
        raise ValueError("case_id 已存在；系统不会静默覆盖旧案例。")
    case_dir.mkdir(parents=True, exist_ok=False)
    documents_dir = case_dir / "documents"
    documents_dir.mkdir()
    try:
        normalized_documents: list[dict[str, Any]] = []
        for document in manifest["documents"]:
            stored_name = f"{document['document_id']}.pdf"
            destination = documents_dir / stored_name
            destination.write_bytes(files[document["source_file"]])
            normalized_documents.append(
                {
                    **document,
                    "original_package_path": document["source_file"],
                    "source_file": Path(document["source_file"]).name,
                    "storage_relpath": _relative_path_within(destination, workspace_root).as_posix(),
                }
            )
        complete_years = sorted(
            {
                row["year"]
                for row in financial_rows
                if row["field_kind"] in {"revenue", "accounts_receivable"}
            },
            reverse=True,
        )
        document_by_id = {item["document_id"]: item for item in normalized_documents}
        for row in financial_rows:
            document = document_by_id[row["document_id"]]
            row.update(
                {
                    "source_file": document["source_file"],
                    "storage_relpath": document["storage_relpath"],
                    "disclosure_date": document["disclosure_date"],
                    "file_sha256": document["sha256"],
                    "source_review_status": "import_validated_pending_human_professional_confirmation",
                }
            )
        case = {
            **manifest,
            "ai_generated_content_notice": AI_GENERATED_CONTENT_NOTICE,
            # 仅由已认证服务端写入；浏览器提交的组织编号不会进入案例清单。
            "tenant_id": normalized_tenant,
            "owner_user_id": owner_user_id,
            "storage_backend": "supabase_private" if normalized_tenant else "local_competition",
            "runtime_materialization": "tenant_scoped_ephemeral" if normalized_tenant else "local_authoritative",
            "documents": normalized_documents,
            "available_years": [year for year in complete_years if year - 1 in complete_years],
            "three_year_r1_ready": len(complete_years) >= 3,
            "source_snapshot_id": f"{manifest['case_id'].lower()}-{_sha256_bytes(content)[:12].lower()}",
            "source_review_status": "import_validated_pending_human_professional_confirmation",
            "legal_sample_confirmation_status": (
                "operator_attested_manifest_record_pending_independent_review"
                if manifest["model_transfer_allowed"]
                else "pending_human_confirmation"
            ),
            "import_confirmation": {
                "authorized_or_public_source_asserted": authorized,
                "desensitization_or_personal_information_review_asserted": desensitized,
                "boundary": "接口勾选和manifest记录用于留痕，不替代独立真人或必要合规人员复核。",
            },
            "registry_mode": "imported_template",
            "structured_evidence": structured_evidence,
            "material_gaps": material_gaps,
            "imported_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
            "package_sha256": _sha256_bytes(content),
            "package_id": f"PKG-{uuid.uuid4().hex[:12].upper()}",
        }
        (case_dir / "financial_fields.json").write_text(
            json.dumps(financial_rows, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        (case_dir / "case.json").write_text(json.dumps(case, ensure_ascii=False, indent=2), encoding="utf-8")
        return case
    except Exception:
        # 仅清理由本次导入刚创建、且已确认位于 cases/<case_id> 的隔离目录。
        for child in sorted(case_dir.rglob("*"), reverse=True):
            if child.is_file():
                child.unlink()
            elif child.is_dir():
                child.rmdir()
        case_dir.rmdir()
        raise


def register_cninfo_case(
    workspace_root: Path,
    *,
    case_id: str,
    company: dict[str, Any],
    documents: list[dict[str, Any]],
    retention_expires_at: str = "2026-12-31",
    tenant_id: str | None = None,
    owner_user_id: str | None = None,
) -> dict[str, Any]:
    """登记已经通过巨潮下载和 PDF 校验的公开年报案例。

    该入口只负责来源登记和文件保存，暂不假设财务字段已经专业确认。
    因此初始 financial_fields.json 为空，RAG 可以先运行，完整分析必须等字段闸门通过。
    """

    # 自动案例沿用现有案例目录协议，因此 RAG、证据回查和旧接口可以复用。
    # 与人工 ZIP 导入不同，这里先登记空的 financial_fields，再由候选提取器补充。
    # 案例编号只允许白名单字符，避免公司名称或外部输入形成路径穿越。
    if not CASE_ID_PATTERN.fullmatch(case_id):
        raise ValueError("巨潮案例编号格式不合法。")
    if not documents:
        raise ValueError("巨潮案例至少需要一份年报。")
    required_company = ("ticker", "company_name", "company_alias", "org_id", "market")
    if any(not str(company.get(key) or "").strip() for key in required_company):
        raise ValueError("巨潮公司元数据不完整。")
    normalized_documents: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    seen_years: set[int] = set()
    fingerprint_rows: list[dict[str, Any]] = []
    # 文档按年度倒序保存，页面默认优先展示最新报告，但不会丢弃历史年度。
    for item in sorted(documents, key=lambda row: int(row.get("report_year", 0)), reverse=True):
        document_id = str(item.get("document_id") or "").strip().upper()
        if not DOCUMENT_ID_PATTERN.fullmatch(document_id) or document_id in seen_ids:
            raise ValueError("巨潮文档编号格式不合法或重复。")
        content = item.get("content")
        if not isinstance(content, bytes) or not content.startswith(b"%PDF-"):
            # 登记层再次检查文件头，防止调用方绕过下载校验直接写入非 PDF。
            raise ValueError(f"{document_id} 不是已经校验的 PDF 内容。")
        try:
            report_year = int(item["report_year"])
        except (KeyError, TypeError, ValueError) as error:
            raise ValueError(f"{document_id} 缺少有效报告年度。") from error
        if report_year in seen_years:
            raise ValueError(f"巨潮案例重复登记 {report_year} 年报告。")
        disclosure_date = _iso_date(item.get("announcement_date"), "公告日期")
        source_url = str(item.get("source_url") or "")
        if not source_url.startswith("https://static.cninfo.com.cn/finalpage/") or not source_url.lower().endswith(".pdf"):
            raise ValueError(f"{document_id} 来源不是巨潮 PDF 原件。")
        actual_hash = _sha256_bytes(content)
        expected_hash = str(item.get("sha256") or actual_hash).upper()
        # 哈希是来源快照的一部分，同一案例不会静默替换不同版本的原件。
        if expected_hash != actual_hash:
            raise ValueError(f"{document_id} 文件 SHA-256 不一致。")
        file_name = f"{document_id}.pdf"
        normalized_documents.append(
            {
                "document_id": document_id,
                "source_file": file_name,
                "original_source_file": str(item.get("source_file") or file_name)[:160],
                "storage_relpath": f"backend/runtime/cases/{case_id}/documents/{file_name}",
                "content_store_relpath": (
                    _relative_path_within(
                        _pdf_content_store_dir(workspace_root) / f"{actual_hash.upper()}.pdf",
                        workspace_root,
                    ).as_posix()
                ),
                "document_type": "annual_report",
                "report_year": report_year,
                "disclosure_date": disclosure_date,
                "announcement_title": str(item.get("announcement_title") or "")[:300],
                "source_url": source_url,
                "sha256": actual_hash,
                "file_sha256": actual_hash,
                "byte_count": int(item.get("byte_count") or len(content)),
                "page_count": int(item.get("page_count") or 0),
                "validation_status": str(item.get("validation_status") or "passed"),
                "source_mode": "cninfo_official",
                "selection_reason": str(item.get("selection_reason") or "")[:300],
                "candidate_count": int(item.get("candidate_count") or 1),
                "candidate_urls": list(item.get("candidate_urls") or [])[:20],
            }
        )
        fingerprint_rows.append(
            {
                "document_id": document_id,
                "report_year": report_year,
                "source_url": source_url,
                "sha256": actual_hash,
            }
        )
        seen_ids.add(document_id)
        seen_years.add(report_year)

    source_snapshot_id = hashlib.sha256(
        json.dumps(fingerprint_rows, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()[:24].lower()
    case_dir = _case_dir(workspace_root, case_id)
    # 测试命名空间和正式运行都可能改变 backend/runtime 的相对位置，
    # 因此来源路径必须根据实际案例目录计算，不能手写未带命名空间的路径。
    for document in normalized_documents:
        document["storage_relpath"] = (
            _relative_path_within(
                case_dir / "documents" / document["source_file"],
                workspace_root,
            ).as_posix()
        )
    if case_dir.exists():
        existing = _wait_for_registered_case(case_dir, source_snapshot_id)
        if existing is not None:
            return existing

    # exists()+mkdir 之间仍可能被另一 worker 抢先创建。失败时等待其写完
    # case.json 并比较快照；相同快照复用，不同快照继续失败关闭。
    while True:
        try:
            case_dir.mkdir(parents=True, exist_ok=False)
            break
        except FileExistsError:
            existing = _wait_for_registered_case(case_dir, source_snapshot_id)
            if existing is not None:
                return existing
    documents_dir = case_dir / "documents"
    documents_dir.mkdir()
    try:
        for item, normalized in zip(
            sorted(documents, key=lambda row: int(row.get("report_year", 0)), reverse=True),
            normalized_documents,
        ):
            destination = documents_dir / normalized["source_file"]
            shared_blob = _ensure_content_addressed_pdf(workspace_root, normalized["sha256"], item["content"])
            try:
                os.link(shared_blob, destination)
            except FileExistsError:
                pass
            except OSError as error:
                raise ValueError("当前文件系统不支持 PDF 内容寻址硬链接，已停止写入以避免重复存储。") from error
        disclosure_dates = [item["disclosure_date"] for item in normalized_documents]
        # 以下元数据明确标出公开来源、待人工确认和模型传输关闭状态。
        case = {
            "schema_version": CASE_SCHEMA_VERSION,
            "case_id": case_id,
            "company_name": str(company["company_name"])[:200],
            "company_alias": str(company["company_alias"])[:100],
            "ticker": str(company["ticker"])[:30],
            "org_id": str(company["org_id"])[:80],
            "market": str(company["market"])[:20],
            "t0": max(disclosure_dates),
            "currency": "CNY",
            "amount_unit": "元",
            "statement_scope": "合并",
            "sample_type": "public",
            "tenant_id": tenant_id,
            "owner_user_id": owner_user_id,
            "storage_backend": "supabase_private" if tenant_id else "public_official_or_local_cache",
            "model_transfer_allowed": False,
            "retention_expires_at": retention_expires_at,
            "legal_sample_confirmation_status": "pending_human_confirmation",
            "source_snapshot_id": source_snapshot_id,
            "source_review_status": "cninfo_download_validated_pending_human_professional_confirmation",
            "documents": normalized_documents,
            "available_report_years": sorted(seen_years, reverse=True),
            "available_years": [],
            "three_year_r1_ready": False,
            "registry_mode": "cninfo_official_auto",
            "structured_evidence": [],
            "material_gaps": ["营业收入和应收账款结构化字段待提取与人工页码确认"],
            "financial_fields_status": "not_extracted",
            "ai_generated_content_notice": AI_GENERATED_CONTENT_NOTICE,
            "registered_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        }
        (case_dir / "financial_fields.json").write_text("[]", encoding="utf-8")
        (case_dir / "case.json").write_text(json.dumps(case, ensure_ascii=False, indent=2), encoding="utf-8")
        return case
    except Exception:
        # 这里只清理由本次函数创建的单个案例目录，绝不触碰其他案例。
        # 清理失败也不影响已经存在的标准案例和其他自动案例。
        for child in sorted(case_dir.rglob("*"), reverse=True):
            if child.is_file():
                child.unlink()
            elif child.is_dir():
                child.rmdir()
        case_dir.rmdir()
        raise


def update_cninfo_financial_fields(
    workspace_root: Path,
    case_id: str,
    rows: list[dict[str, Any]],
    *,
    status: str,
    material_gaps: list[str] | None = None,
    specialized_required_fields: list[str] | None = None,
    industry_rule: str | None = None,
) -> dict[str, Any]:
    """把字段候选写入巨潮案例，并重新计算可用于 R1 的连续年度。"""

    # 更新只允许发生在自动注册案例，人工 ZIP 案例仍遵循原有财务字段协议。
    case = get_case(workspace_root, case_id)
    # ZIP 案例仍沿用原有人工字段协议，不能通过新接口改变其审查方式。
    if case is None or case.get("registry_mode") != "cninfo_official_auto":
        raise ValueError("不是可更新的巨潮自动案例。")
    # 自动候选更新必须读取旧记录，保证同一来源重新跑时不会丢失真人复核。
    documents = {item["document_id"]: item for item in case.get("documents", [])}
    # 候选快照和规则输入值分开保存，修正只能改变后者。
    previous_path = _case_dir(workspace_root, case_id) / "financial_fields.json"
    try:
        previous_rows = json.loads(previous_path.read_text(encoding="utf-8")) if previous_path.is_file() else []
    except (OSError, json.JSONDecodeError):
        previous_rows = []
    previous_by_field_id = {
        str(row.get("field_id") or f"{row.get('field_kind')}_{row.get('year')}"): row
        for row in previous_rows
        if isinstance(row, dict)
    }
    normalized: list[dict[str, Any]] = []
    seen: set[tuple[str, int]] = set()
    # 同一字段年度只能有一个候选，避免重复候选让规则误以为年度已完整。
    for row in rows:
        kind = str(row.get("field_kind") or "")
        year = int(row.get("year"))
        document_id = str(row.get("document_id") or "").upper()
        if kind not in ALLOWED_FIELD_KINDS or document_id not in documents:
            raise ValueError("巨潮自动字段包含未知类型或未登记文档。")
        if (kind, year) in seen:
            raise ValueError(f"巨潮自动字段重复：{kind}/{year}")
        seen.add((kind, year))
        document = documents[document_id]
        if document["report_year"] != year:
            raise ValueError(f"{kind}/{year} 未绑定同年度年报。")
        value = float(row["value"])
        pdf_page = int(row["pdf_page"])
        if pdf_page < 1 or not math.isfinite(value):
            raise ValueError("巨潮自动字段页码或金额无效。")
        evidence_id = str(row.get("evidence_id") or f"{case_id}_{FIELD_PREFIX[kind]}_{year}").upper()
        if not DOCUMENT_ID_PATTERN.fullmatch(evidence_id):
            raise ValueError("巨潮自动字段 evidence_id 不合法。")
        # 字段编号由字段种类和年度组成，页面、接口和审计日志共同使用。
        field_id = f"{kind}_{year}"
        previous = previous_by_field_id.get(field_id)
        # 只有文档、页码和金额都没有变化时，才继承上一轮真人决定。
        previous_value = None
        if previous:
            try:
                previous_value = float(previous.get("candidate", {}).get("value", previous.get("value")))
            except (TypeError, ValueError):
                previous_value = None
        same_candidate = bool(
            previous
            and previous_value is not None
            and math.isfinite(previous_value)
            and previous_value == value
            and int(previous.get("candidate", {}).get("pdf_page", previous.get("pdf_page"))) == pdf_page
            and str(previous.get("document_id")) == document_id
        )
        # 来源发生变化时必须重新待确认，不能把旧年报的签字带到新版本。
        human_review = (
            deepcopy(previous.get("human_review"))
            if same_candidate and isinstance(previous.get("human_review"), dict)
            else {"status": "pending", "decision": None}
        )
        # 历史记录追加保存，便于评审回看每次确认、修正和拒绝。
        human_review_history = (
            deepcopy(previous.get("human_review_history"))
            if same_candidate and isinstance(previous.get("human_review_history"), list)
            else []
        )
        normalized.append(
            {
                "case_id": case_id,
                "evidence_id": evidence_id,
                "field_id": field_id,
                "field_kind": kind,
                "year": year,
                "value": value,
                "candidate": {
                    "value": value,
                    "pdf_page": pdf_page,
                    "document_id": document_id,
                    "locator": str(row.get("locator") or "自动提取候选；待人工回页确认")[:300],
                },
                "unit": str(row.get("unit") or ("%" if kind in RATIO_FIELD_KINDS else case.get("amount_unit", "元"))),
                "source_unit": str(row.get("source_unit") or row.get("unit") or ""),
                "currency": case.get("currency", "CNY"),
                "statement_scope": case.get("statement_scope", "合并"),
                "field_basis": str(row.get("field_basis") or ("gross" if kind == "accounts_receivable" else "reported")),
                "document_id": document_id,
                "source_file": document["source_file"],
                "storage_relpath": document["storage_relpath"],
                "disclosure_date": document["disclosure_date"],
                "announcement_title": document["announcement_title"],
                "source_url": document["source_url"],
                "pdf_page": pdf_page,
                "print_page": row.get("print_page"),
                "locator": str(row.get("locator") or "自动提取候选；待人工回页确认")[:300],
                "file_sha256": document["sha256"],
                "source_review_status": str(row.get("source_review_status") or "auto_extracted_pending_human_page_confirmation"),
                "extraction_method": str(row.get("extraction_method") or "pdf_text_heuristic_candidate")[:100],
                "raw_excerpt": str(row.get("raw_excerpt") or "")[:1000],
                "human_review": human_review,
                "human_review_history": human_review_history,
            }
        )
    # R1 的可用年度必须同时有营业收入和应收账款，不能用单个字段的最长年度误判完整性。
    # 行业专用规则不一定使用 R1 字段；先分别计算普通字段和专用字段，避免银行、保险、
    # 券商案例明明有候选却因为 R1 为空而在页面显示“无可用年度”。
    complete_years = _complete_r1_years(normalized)
    r1_available_years = [year for year in complete_years if year - 1 in complete_years]
    confirmed_years = _complete_r1_years(normalized, accepted_only=True)
    r1_human_confirmed_available_years = [year for year in confirmed_years if year - 1 in confirmed_years]
    # available_years 表示技术上有候选，不代表已经可以进入规则。
    case["available_years"] = r1_available_years
    case["r1_available_years"] = r1_available_years
    if specialized_required_fields:
        specialized_complete_years = sorted(
            year
            for year in {int(row["year"]) for row in normalized}
            if all(
                any(item.get("field_kind") == kind and int(item.get("year")) == year for item in normalized)
                for kind in specialized_required_fields
            )
        )
        specialized_available_years = [
            year for year in reversed(specialized_complete_years) if year - 1 in specialized_complete_years
        ]
        case["specialized_required_fields"] = list(specialized_required_fields)
        case["specialized_rule"] = industry_rule
        case["industry_rule_version"] = INDUSTRY_RULES_VERSION if industry_rule else None
        case["specialized_available_years"] = specialized_available_years
        case["available_years"] = specialized_available_years
        confirmed_specialized_years = sorted(
            year
            for year in {int(row["year"]) for row in normalized}
            if all(
                any(
                    item.get("field_kind") == kind
                    and int(item.get("year")) == year
                    and _cninfo_field_is_human_accepted(item)
                    for item in normalized
                )
                for kind in specialized_required_fields
            )
        )
        case["specialized_human_confirmed_available_years"] = [
            year for year in reversed(confirmed_specialized_years) if year - 1 in confirmed_specialized_years
        ]
    else:
        # 案例从专用规则切回普通 R1/R2 时，不能继续沿用上一轮专用年度和规则键。
        # 否则下一次页面读取会把旧行业结果误当成本轮字段完整性。
        for key in (
            "specialized_required_fields",
            "specialized_rule",
            "industry_rule_version",
            "specialized_available_years",
            "specialized_human_confirmed_available_years",
        ):
            case.pop(key, None)
    # human_confirmed_available_years 才是人工闸门之后允许规则使用的期间。
    case["human_confirmed_available_years"] = (
        case.get("specialized_human_confirmed_available_years", r1_human_confirmed_available_years)
        if specialized_required_fields
        else r1_human_confirmed_available_years
    )
    case["three_year_r1_ready"] = _has_three_consecutive_years(complete_years)
    case["human_confirmed_three_year_r1_ready"] = _has_three_consecutive_years(confirmed_years)
    case["financial_fields_status"] = status
    case["material_gaps"] = list(material_gaps or [])
    case["source_review_status"] = (
        # 即使技术上通过，页面和口径仍需真人回到原 PDF 确认。
        "cninfo_fields_candidate_pending_human_professional_confirmation"
        if normalized
        else "cninfo_download_validated_pending_human_professional_confirmation"
    )
    case["updated_at"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    case_dir = _case_dir(workspace_root, case_id)
    (case_dir / "financial_fields.json").write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")
    (case_dir / "case.json").write_text(json.dumps(case, ensure_ascii=False, indent=2), encoding="utf-8")
    return case


def _cninfo_field_is_human_accepted(row: dict[str, Any]) -> bool:
    """只有真人确认或真人修正后的字段才能进入规则输入。"""

    # 拒绝、待确认和缺少复核字段都必须保持关闭，不以页面按钮状态推断。
    return row.get("human_review", {}).get("decision") in {"confirm", "correct"}


def _complete_r1_years(rows: list[dict[str, Any]], *, accepted_only: bool = False) -> list[int]:
    """只返回收入和应收账款同时存在的年度，必要时再叠加人工确认条件。"""

    complete: list[int] = []
    for year in sorted({int(row["year"]) for row in rows}, reverse=True):
        kinds = {
            row.get("field_kind")
            for row in rows
            if int(row.get("year")) == year
            and (not accepted_only or _cninfo_field_is_human_accepted(row))
        }
        if {"revenue", "accounts_receivable"}.issubset(kinds):
            complete.append(year)
    return complete


def _has_three_consecutive_years(years: list[int]) -> bool:
    """三年准备状态必须是连续年度，而不是任意三个离散年度。"""

    values = set(years)
    return any(year - 1 in values and year - 2 in values for year in values)


def _recompute_cninfo_human_status(case: dict[str, Any], rows: list[dict[str, Any]]) -> None:
    """更新字段候选的真人状态，但不改写原始自动候选和历史复核记录。"""

    technical_years = _complete_r1_years(rows)
    confirmed_years = _complete_r1_years(rows, accepted_only=True)
    case["available_years"] = [year for year in technical_years if year - 1 in technical_years]
    case["human_confirmed_available_years"] = [year for year in confirmed_years if year - 1 in confirmed_years]
    case["three_year_r1_ready"] = _has_three_consecutive_years(technical_years)
    case["human_confirmed_three_year_r1_ready"] = _has_three_consecutive_years(confirmed_years)
    required_rows = [
        row
        for row in rows
        if row.get("field_kind") in {"revenue", "accounts_receivable"}
    ]
    accepted_required = bool(required_rows) and all(_cninfo_field_is_human_accepted(row) for row in required_rows)
    case["financial_fields_status"] = "human_confirmed" if accepted_required else "pending_human_review"
    case["source_review_status"] = (
        "cninfo_fields_human_confirmed"
        if accepted_required
        else "cninfo_fields_candidate_pending_human_professional_confirmation"
    )


def confirm_cninfo_field(
    workspace_root: Path,
    case_id: str,
    confirmation: dict[str, Any],
) -> dict[str, Any]:
    """追加保存一条真人字段处理记录，并返回当前字段和案例闸门状态。"""

    case = get_case(workspace_root, case_id)
    if case is None or case.get("registry_mode") != "cninfo_official_auto":
        raise ValueError("只有巨潮自动案例支持字段真人确认。")
    # 复核人由操作者明确填写，AI 不代填姓名、角色或签字日期。
    reviewer = str(confirmation.get("reviewer") or "").strip()
    decision = str(confirmation.get("decision") or "")
    reason = str(confirmation.get("reason") or "").strip()
    # 空白复核人不是有效的人工作业记录，接口必须拒绝保存。
    if not reviewer:
        raise ValueError("字段确认必须填写真实复核人或团队角色。")
    if decision not in {"confirm", "correct", "reject"}:
        raise ValueError("字段处理动作只能是 confirm、correct 或 reject。")
    # 修正和拒绝会改变后续规则资格，必须留下可解释的原因。
    if decision in {"correct", "reject"} and not reason:
        raise ValueError("修正或拒绝字段必须填写原因。")
    path = _case_dir(workspace_root, case_id) / "financial_fields.json"
    if not path.is_file():
        raise ValueError("案例尚未形成字段候选。")
    rows = json.loads(path.read_text(encoding="utf-8"))
    field_id = str(confirmation.get("field_id") or "")
    # 兼容字段确认功能上线前已经生成的自动案例，旧记录按字段种类和年度补出编号。
    row = next(
        (
            item
            for item in rows
            if str(item.get("field_id") or f"{item.get('field_kind')}_{item.get('year')}") == field_id
        ),
        None,
    )
    # 字段编号找不到时停止，不根据年度和字段名称模糊匹配其他候选。
    if row is None:
        raise ValueError("未找到对应字段候选。")
    row["field_id"] = field_id
    # 旧记录没有候选快照时，以原始自动值建立一次不可变基线。
    row.setdefault(
        "candidate",
        {
            "value": row.get("value"),
            "pdf_page": row.get("pdf_page"),
            "document_id": row.get("document_id"),
            "locator": row.get("locator"),
        },
    )
    row.setdefault("human_review_history", [])
    documents = {item["document_id"]: item for item in case.get("documents", [])}
    # original 保存处理前的金额、页码和定位，修正时不能覆盖证据链。
    original = {
        "value": row.get("value"),
        "pdf_page": row.get("pdf_page"),
        "document_id": row.get("document_id"),
        "locator": row.get("locator"),
    }
    # 修正动作仍绑定原登记文档，并检查页码不超过已校验 PDF。
    if decision == "correct":
        corrected_value = confirmation.get("corrected_value")
        corrected_page = confirmation.get("corrected_pdf_page")
        if corrected_value is None or corrected_page is None:
            raise ValueError("修正字段必须同时提供金额和 PDF 页码。")
        try:
            normalized_value = float(corrected_value)
            normalized_page = int(corrected_page)
        except (TypeError, ValueError) as error:
            raise ValueError("修正后的金额或 PDF 页码不是有效数值。") from error
        if not math.isfinite(normalized_value) or normalized_page < 1:
            raise ValueError("修正后的金额必须为有限数值，PDF 页码必须大于零。")
        document = documents.get(str(row.get("document_id")))
        if document and normalized_page > int(document.get("page_count") or 0):
            raise ValueError("修正后的 PDF 页码超过已校验原件页数。")
        row["value"] = normalized_value
        row["pdf_page"] = normalized_page
        if confirmation.get("corrected_locator"):
            row["locator"] = str(confirmation["corrected_locator"])[:300]
        row["source_review_status"] = "human_corrected"
    elif decision == "confirm":
        row["source_review_status"] = "human_confirmed"
    else:
        row["source_review_status"] = "human_rejected"
    # 服务器生成复核时间，防止浏览器伪造历史时间线。
    review = {
        "status": {"confirm": "confirmed", "correct": "corrected", "reject": "rejected"}[decision],
        "decision": decision,
        "reviewer": reviewer,
        "reviewed_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "reason": reason,
        "original": original,
    }
    # 当前决定供规则闸门读取，历史列表供项目审查和答辩回查。
    row.setdefault("human_review_history", []).append(deepcopy(review))
    row["human_review"] = review
    # 每次单字段处理后重新计算整个案例闸门，避免页面局部状态失真。
    _recompute_cninfo_human_status(case, rows)
    case["updated_at"] = time.strftime("%Y-%m-%dT%H:%M:%S%z")
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    (_case_dir(workspace_root, case_id) / "case.json").write_text(json.dumps(case, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"case": case, "field": row, "readiness": get_cninfo_field_readiness(workspace_root, case_id, ["R1", "R2"], max(case.get("available_years") or [0]))}


def get_cninfo_field_readiness(
    workspace_root: Path,
    case_id: str,
    rule_ids: Iterable[str],
    current_year: int,
) -> list[str]:
    """按规则检查巨潮字段是否已经过真人确认，返回阻断原因。"""

    case = get_case(workspace_root, case_id)
    if case is None or case.get("registry_mode") != "cninfo_official_auto":
        return []
    # readiness 与技术候选数量分开计算，明确回答“能不能进规则”。
    rows = get_financial_rows(workspace_root, case_id)
    row_map = {(row.get("field_kind"), int(row.get("year"))): row for row in rows}
    required: set[tuple[str, int]] = set()
    # R1 只要求收入和应收主字段，三年候选存在时也必须确认被实际使用的第三年。
    if "R1" in set(rule_ids):
        required.update({("revenue", current_year), ("revenue", current_year - 1), ("accounts_receivable", current_year), ("accounts_receivable", current_year - 1)})
        if all((kind, current_year - 2) in row_map for kind in ("revenue", "accounts_receivable")):
            required.update({("revenue", current_year - 2), ("accounts_receivable", current_year - 2)})
    # R2 额外要求经营现金流，净利润存在时不能跳过其人工回查。
    if "R2" in set(rule_ids):
        required.update({("revenue", current_year), ("revenue", current_year - 1), ("operating_cash_flow", current_year), ("operating_cash_flow", current_year - 1)})
        if ("net_profit", current_year) in row_map:
            required.add(("net_profit", current_year))
    issues: list[str] = []
    # 按年度和字段稳定排序，确保页面提示和日志在多次运行中可复现。
    for kind, year in sorted(required, key=lambda item: (item[1], item[0])):
        row = row_map.get((kind, year))
        if row is None:
            issues.append(f"{year}年{kind}字段缺失")
        elif not _cninfo_field_is_human_accepted(row):
            issues.append(f"{year}年{kind}字段尚未真人确认")
    return issues


def _standard_financial_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for evidence_id, evidence in standard_data.EVIDENCE.items():
        document_id = f"STD-AR-{evidence['year']}-{evidence['file_sha256'][:12]}"
        rows.append(
            {
                "case_id": standard_data.CASE_ID,
                "evidence_id": evidence_id,
                "field_kind": evidence["field_kind"],
                "year": evidence["year"],
                "value": evidence["value"],
                "unit": evidence["unit"],
                "currency": "CNY",
                "statement_scope": "合并",
                "field_basis": "net" if evidence["field_kind"] == "accounts_receivable" else "reported",
                "document_id": document_id,
                "source_file": evidence["source_file"],
                "storage_relpath": evidence["source_file"],
                "disclosure_date": evidence["disclosure_date"],
                "announcement_title": evidence["announcement_title"],
                "source_url": evidence["source_url"],
                "pdf_page": evidence["pdf_page"],
                "print_page": evidence["print_page"],
                "locator": evidence["locator"],
                "file_sha256": evidence["file_sha256"],
                "source_review_status": standard_data.SOURCE_REVIEW_STATUS,
            }
        )
    return rows


def get_financial_rows(
    workspace_root: Path,
    case_id: str,
    *,
    tenant_id: str | None = None,
) -> list[dict[str, Any]]:
    """读取结构化字段；租户参数必须与案例临时目录和清单归属同时匹配。"""

    normalized_tenant = _normalized_server_tenant_id(tenant_id)
    if case_id == standard_data.CASE_ID and normalized_tenant is None:
        return _standard_financial_rows()
    case = get_case(workspace_root, case_id, tenant_id=normalized_tenant)
    if case is None:
        raise KeyError(case_id)
    path = _case_dir(workspace_root, case_id, tenant_id=normalized_tenant) / "financial_fields.json"
    return json.loads(path.read_text(encoding="utf-8"))


def get_case_documents(
    workspace_root: Path,
    case_id: str,
    *,
    tenant_id: str | None = None,
) -> list[dict[str, Any]]:
    """返回当前作用域文档登记副本，不跨租户回退同名案例。"""

    case = get_case(workspace_root, case_id, tenant_id=tenant_id)
    if case is None:
        raise KeyError(case_id)
    return deepcopy(case["documents"])


def resolve_case_document(
    workspace_root: Path,
    case_id: str,
    document_id: str,
    *,
    tenant_id: str | None = None,
) -> tuple[Path, dict[str, Any]] | None:
    """解析已登记原件；非内置案例必须落在当前作用域的 documents 目录。"""

    normalized_tenant = _normalized_server_tenant_id(tenant_id)
    case = get_case(workspace_root, case_id, tenant_id=normalized_tenant)
    if case is None:
        return None
    document = next((item for item in case["documents"] if item["document_id"] == document_id), None)
    if document is None:
        return None
    path = (workspace_root / document["storage_relpath"]).resolve()
    if not path.is_file():
        return None
    allowed_root = (
        workspace_root.resolve()
        if case_id == standard_data.CASE_ID and normalized_tenant is None
        else (_case_dir(workspace_root, case_id, tenant_id=normalized_tenant) / "documents").resolve()
    )
    if not _path_is_within(path, allowed_root):
        return None
    # 导入和巨潮登记都以 document_id 作为不可变文件名；拒绝元数据把同租户
    # 另一文档甚至另一案例的合法路径冒充成本次登记原件。
    if (
        not (case_id == standard_data.CASE_ID and normalized_tenant is None)
        and _canonical_path_text(path) != _canonical_path_text(allowed_root / f"{document_id}.pdf")
    ):
        return None
    return path, deepcopy(document)


def _select(rows: Iterable[dict[str, Any]], kind: str, year: int) -> dict[str, Any] | None:
    return next((deepcopy(row) for row in rows if row["field_kind"] == kind and row["year"] == year), None)


_SOURCE_UNIT_MULTIPLIERS = {
    "元": 1.0,
    "千元": 1_000.0,
    "万元": 10_000.0,
    "百万元": 1_000_000.0,
    "亿元": 100_000_000.0,
    "%": 1.0,
}
_RATIO_FIELD_KINDS = {
    "nonperforming_loan_ratio",
    "provision_coverage_ratio",
}
_HUMAN_ACCEPTED_REVIEW_STATUSES = {
    "human_confirmed",
    "human_corrected",
    "owner_confirmed_registered_public_evidence",
}


def financial_field_candidate_quality_issues(row: dict[str, Any]) -> list[str]:
    """识别不能直接进入计算的自动候选；真人确认或更正可显式解除闸门。"""

    review_status = str(row.get("source_review_status") or "").strip()
    if review_status in _HUMAN_ACCEPTED_REVIEW_STATUSES:
        return []
    existing = [str(item) for item in row.get("candidate_quality_issues") or [] if str(item).strip()]
    extraction_method = str(row.get("extraction_method") or "")
    is_automatic_candidate = "heuristic" in extraction_method or "pending" in review_status
    if not is_automatic_candidate:
        return list(dict.fromkeys(existing))

    issues = list(existing)
    field_kind = str(row.get("field_kind") or "")
    source_unit = str(row.get("source_unit") or row.get("unit") or "")
    try:
        value = float(row.get("value"))
        source_value = value / _SOURCE_UNIT_MULTIPLIERS[source_unit]
    except (KeyError, TypeError, ValueError, ZeroDivisionError):
        source_value = None
    excerpt = re.sub(r"\s+", "", str(row.get("raw_excerpt") or row.get("excerpt") or ""))

    is_ratio = field_kind in _RATIO_FIELD_KINDS or str(row.get("unit") or "") == "%"
    if is_ratio:
        if source_unit != "%":
            issues.append("比例字段的来源单位不是百分比，自动候选不能进入计算。")
        suspicious_ratio_context = bool(
            re.search(r"(?:拨备覆盖率|不良贷款率)\s*[（(]\d{1,3}[）)]", excerpt)
            or re.search(r"(?:拨备覆盖率|不良贷款率)\s*\d{1,2}(?=\D|$)", excerpt)
            or re.search(r"(?:≥|>=)\s*\d+(?:\.\d+)?\s*[（(]?注", excerpt)
            or re.search(r"\d+(?:\.\d+)?\s*(?:天|个月)", excerpt)
        )
        if suspicious_ratio_context:
            issues.append("比例候选疑似取到附注编号、监管阈值或期限，须人工回页确认。")
    elif source_value is not None and abs(source_value) <= 100:
        issues.append(
            f"自动提取金额的原始值仅为 {source_value:g}{source_unit or '（单位未知）'}，"
            "疑似附注号、序号或叙述数字，须人工回页确认。"
        )
    return list(dict.fromkeys(issues))


def annotate_financial_field_quality(row: dict[str, Any]) -> dict[str, Any]:
    """返回带候选质量状态的副本，保留原始值供人工复核。"""

    annotated = deepcopy(row)
    issues = financial_field_candidate_quality_issues(annotated)
    annotated["candidate_quality_issues"] = issues
    annotated["candidate_quality_status"] = (
        "blocked_pending_human_confirmation" if issues else "passed_technical_candidate"
    )
    return annotated


def annotate_financial_field_rows_quality(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    """在逐行闸门之上检查同字段连续年度的异常数量级变化。"""

    annotated = [annotate_financial_field_quality(row) for row in rows]
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in annotated:
        grouped.setdefault(str(row.get("field_kind") or ""), []).append(row)
    for field_rows in grouped.values():
        ordered = sorted(field_rows, key=lambda item: int(item.get("year") or 0))
        for previous, current in zip(ordered, ordered[1:]):
            if int(current.get("year") or 0) - int(previous.get("year") or 0) != 1:
                continue
            # 数量级闸门只用于 PDF 启发式自动候选。内置、导入或已结构化
            # 字段即使尚待人工复核，也可能存在合法的现金流变号或并购跃升；
            # 不能仅凭“pending”状态把它们改写成自动抽取错位。
            if not all(
                "heuristic" in str(row.get("extraction_method") or "")
                for row in (previous, current)
            ):
                continue
            try:
                previous_value = abs(float(previous.get("value")))
                current_value = abs(float(current.get("value")))
            except (TypeError, ValueError):
                continue
            smaller = min(previous_value, current_value)
            larger = max(previous_value, current_value)
            if smaller <= 0 or larger / smaller < 10:
                continue
            issue = (
                f"同字段连续年度金额相差 {larger / smaller:.1f} 倍，"
                "可能存在错列、单位或表内子项误取，须人工回页确认。"
            )
            for row in (previous, current):
                if str(row.get("source_review_status") or "") in _HUMAN_ACCEPTED_REVIEW_STATUSES:
                    continue
                row["candidate_quality_issues"] = list(
                    dict.fromkeys(list(row.get("candidate_quality_issues") or []) + [issue])
                )
                row["candidate_quality_status"] = "blocked_pending_human_confirmation"
    return annotated


def calculation_ready_financial_rows(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    """只返回未被拒绝且通过候选质量闸门的字段行。"""

    ready: list[dict[str, Any]] = []
    for row in annotate_financial_field_rows_quality(rows):
        if str(row.get("source_review_status") or "") == "human_rejected":
            continue
        if row["candidate_quality_issues"]:
            continue
        ready.append(row)
    return ready


def _technical_complete_years(rows: list[dict[str, Any]], kinds: tuple[str, ...]) -> list[int]:
    """返回同时具备指定字段候选的年度；不要求真人确认，供公开预筛选年。"""

    # 这里读取的是技术候选，不读取真人决定，避免预筛被人工按钮状态卡住。
    # 每个年度必须同时具备规则所需字段，不能用单个字段拼出伪完整年度。
    years = {int(row["year"]) for row in rows if row.get("field_kind") in kinds}
    return sorted(
        [year for year in years if all(_select(rows, kind, year) is not None for kind in kinds)],
        reverse=True,
    )


def get_cninfo_prescreen_plan(
    workspace_root: Path,
    case_id: str,
    requested_current_year: int,
    rule_ids: Iterable[str] = ("R1",),
) -> dict[str, Any]:
    """为公开财报预筛选择最近可比期间，并记录可运行与缺失规则。"""

    # 公开预筛优先保证审计师先看到结果，再把证据缺口放进结果摘要。
    # 选择年度时只使用真实候选，不对缺失金额、单位或口径作推断。
    case = get_case(workspace_root, case_id)
    if case is None:
        raise KeyError(case_id)
    candidate_rows = get_financial_rows(workspace_root, case_id)
    rows = calculation_ready_financial_rows(candidate_rows)
    annotated_rows = annotate_financial_field_rows_quality(candidate_rows)
    blocked_rows = [row for row in annotated_rows if row.get("candidate_quality_issues")]
    report_years = sorted(
        {int(item.get("report_year")) for item in case.get("documents", []) if item.get("report_year") is not None},
        reverse=True,
    )
    requested = tuple(dict.fromkeys(str(rule_id) for rule_id in rule_ids))
    required_by_rule = {
        "R1": ("revenue", "accounts_receivable"),
        "R2": ("revenue", "operating_cash_flow"),
    }
    plans: dict[str, dict[str, Any]] = {}
    skipped: list[dict[str, Any]] = []
    missing_fields: list[str] = []
    for rule_id in requested:
        kinds = required_by_rule.get(rule_id, ())
        complete_years = _technical_complete_years(rows, kinds)
        pair_years = [year for year in complete_years if year - 1 in complete_years]
        eligible = [year for year in pair_years if year <= requested_current_year]
        # 请求时点之后的公告只属于未来信息，不能因为当前期间缺失就回退使用。
        selected = max(eligible, default=None)
        if selected is None:
            skipped.append(
                {
                    "rule_id": rule_id,
                    "reason": (
                        "可用期间均晚于请求时点，不能使用未来年度。"
                        if pair_years and not eligible
                        else "没有连续两年完整字段，无法进行同比计算。"
                    ),
                    "required_fields": list(kinds),
                }
            )
        else:
            prior_year = selected - 2
            plans[rule_id] = {
                "status": "ready",
                "current_year": selected,
                "previous_year": selected - 1,
                "prior_year": prior_year if prior_year in complete_years else None,
                "complete_years": complete_years,
                "three_year_available": prior_year in complete_years,
            }
        # 对本次下载的每个报告年度登记缺口，让用户知道哪一年不能参与趋势比较。
        # 缺口清单同时服务网页展示、资料索取和后续补充证据续分析。
        for year in report_years:
            if year > requested_current_year:
                continue
            for kind in kinds:
                if _select(rows, kind, year) is None:
                    missing_fields.append(f"{year}年{kind}")

    primary_rule = "R1" if "R1" in plans else next(iter(plans), None)
    selected_plan = plans.get(primary_rule or "")
    selected_year = selected_plan["current_year"] if selected_plan else None
    return {
        "mode": "public_prescreen",
        "requested_current_year": requested_current_year,
        "analysis_current_year": selected_year,
        "analysis_previous_year": selected_year - 1 if selected_year is not None else None,
        "analysis_years": [selected_year, selected_year - 1] if selected_year is not None else [],
        "rule_plans": plans,
        "skipped_rules": skipped,
        "missing_fields": list(dict.fromkeys(missing_fields)),
        "has_calculable_rule": bool(plans),
        "source_candidate_count": len(candidate_rows),
        "usable_source_candidate_count": len(rows),
        "blocked_candidate_count": len(blocked_rows),
        "candidate_quality_issues": list(
            dict.fromkeys(
                issue
                for row in blocked_rows
                for issue in row.get("candidate_quality_issues", [])
            )
        ),
        "human_confirmation": "recommended_before_formal_adoption_or_export",
        "confidence": "technical_candidate_pending_optional_human_confirmation" if rows else "insufficient_data",
    }


def get_period_sources(
    workspace_root: Path,
    case_id: str,
    current_year: int,
    rule_ids: tuple[str, ...] = ("R1",),
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    case = get_case(workspace_root, case_id)
    if case is None:
        raise KeyError(case_id)
    candidate_rows = get_financial_rows(workspace_root, case_id)
    rows = calculation_ready_financial_rows(candidate_rows)
    requested_current_year = current_year
    prescreen_plan = None
    no_calculable_public_period = False
    if case.get("registry_mode") == "cninfo_official_auto":
        prescreen_plan = get_cninfo_prescreen_plan(workspace_root, case_id, current_year, rule_ids)
        if prescreen_plan.get("analysis_current_year") is None:
            # 保留请求年度作为上下文标签，但不构造任何未来或伪造可比字段。
            no_calculable_public_period = True
        else:
            current_year = int(prescreen_plan["analysis_current_year"])
    elif current_year not in case["available_years"]:
        raise KeyError(current_year)
    previous_year = current_year - 1
    prior_year = current_year - 2
    requested: list[tuple[str, str, str, int]] = []
    if "R1" in rule_ids and not no_calculable_public_period:
        requested.extend(
            [
                ("revenue_current", "本年营业收入", "revenue", current_year),
                ("revenue_previous", "上年营业收入", "revenue", previous_year),
                ("ar_current", "本年应收账款", "accounts_receivable", current_year),
                ("ar_previous", "上年应收账款", "accounts_receivable", previous_year),
            ]
        )
        if _select(rows, "revenue", prior_year) and _select(rows, "accounts_receivable", prior_year):
            requested.extend(
                [
                    ("revenue_prior", "前两年营业收入", "revenue", prior_year),
                    ("ar_prior", "前两年应收账款", "accounts_receivable", prior_year),
                ]
            )
        # 账面余额是 R1 增长比较的唯一主口径；准备与净额只作为解释和勾稽证据，
        # 不得静默替代主口径。某一公开年报未登记增强字段时，基础 R1 仍可运行。
        optional_r1_fields = (
            ("ar_allowance_current", "本年应收账款坏账准备", "accounts_receivable_allowance", current_year),
            ("ar_allowance_previous", "上年应收账款坏账准备", "accounts_receivable_allowance", previous_year),
            ("ar_net_current", "本年应收账款净额", "accounts_receivable_net", current_year),
            ("ar_net_previous", "上年应收账款净额", "accounts_receivable_net", previous_year),
        )
        requested.extend(item for item in optional_r1_fields if _select(rows, item[2], item[3]))
    if "R2" in rule_ids and not no_calculable_public_period:
        requested.extend(
            [
                ("revenue_current", "本年营业收入", "revenue", current_year),
                ("revenue_previous", "上年营业收入", "revenue", previous_year),
                ("operating_cash_flow_current", "本年经营活动现金流量净额", "operating_cash_flow", current_year),
                ("operating_cash_flow_previous", "上年经营活动现金流量净额", "operating_cash_flow", previous_year),
            ]
        )
        if _select(rows, "net_profit", current_year):
            requested.append(("net_profit_current", "本年净利润（R2增强项）", "net_profit", current_year))
    sources: list[dict[str, Any]] = []
    seen: set[str] = set()
    for field_id, label, kind, year in requested:
        if field_id in seen:
            continue
        source = _select(rows, kind, year)
        if source is None:
            if kind in {
                "accounts_receivable_allowance",
                "accounts_receivable_net",
                "operating_cash_flow",
                "net_profit",
            }:
                continue
            # 公开预筛允许规则返回 DATA_GAP；不能因为某一字段缺失而阻断其他规则和 RAG。
            if prescreen_plan is not None:
                continue
            raise KeyError(f"{kind}/{year}")
        source.update({"field_id": field_id, "field_label": label})
        sources.append(source)
        seen.add(field_id)
    context = {
        "case_id": case_id,
        "company_name": case["company_name"],
        "company_alias": case.get("company_alias", ""),
        "ticker": case.get("ticker", ""),
        "current_year": current_year,
        "previous_year": previous_year,
        "prior_year": prior_year if any(row["field_id"].endswith("_prior") for row in sources) else None,
        "t0": case["t0"],
        "currency": case["currency"],
        "amount_unit": case["amount_unit"],
        "statement_scope": case["statement_scope"],
        "sample_type": case["sample_type"],
        "model_transfer_allowed": case["model_transfer_allowed"],
        "source_snapshot_id": case["source_snapshot_id"],
        "source_review_status": case["source_review_status"],
        "three_year_r1_ready": case["three_year_r1_ready"],
        "requested_current_year": requested_current_year,
        "analysis_cutoff_year": None if no_calculable_public_period else current_year,
        "public_prescreen": prescreen_plan is not None,
        "prescreen_plan": prescreen_plan,
        "case_evidence_count": len(case.get("structured_evidence", [])),
        "case_material_gaps": list(
            dict.fromkeys(
                list(deepcopy(case.get("material_gaps", [])))
                + list((prescreen_plan or {}).get("candidate_quality_issues") or [])
            )
        ),
    }
    return context, sources


def registered_page_metadata(
    workspace_root: Path, case_id: str, document_id: str, pdf_page: int
) -> dict[str, Any]:
    rows = get_financial_rows(workspace_root, case_id)
    matched = [row for row in rows if row.get("document_id") == document_id and row.get("pdf_page") == pdf_page]
    print_pages = {row.get("print_page") for row in matched if row.get("print_page")}
    return {
        "print_page": next(iter(print_pages)) if len(print_pages) == 1 else None,
        "linked_field_evidence_ids": [row["evidence_id"] for row in matched],
    }


def _placeholder_pdf(year: int) -> bytes:
    document = fitz.open()
    page = document.new_page(width=595, height=842)
    page.insert_text((72, 96), "AuditTrace synthetic template", fontsize=18)
    page.insert_text((72, 130), f"Synthetic annual-report placeholder for {year}.", fontsize=11)
    page.insert_text((72, 154), "Replace with an authorized public or de-identified PDF before formal use.", fontsize=9)
    content = document.tobytes(garbage=4, deflate=True)
    document.close()
    return content


def build_case_template_zip() -> bytes:
    """生成可直接导入的三年合成模板；其模型传输许可默认关闭。"""
    pdfs = {year: _placeholder_pdf(year) for year in (2022, 2023, 2024)}
    documents = []
    for year, content in pdfs.items():
        documents.append(
            {
                "document_id": f"SYNTH-AR-{year}",
                "source_file": f"documents/synthetic_annual_report_{year}.pdf",
                "document_type": "synthetic_annual_report",
                "report_year": year,
                "disclosure_date": f"{year + 1}-04-30",
                "source_url": "",
                "sha256": _sha256_bytes(content),
            }
        )
    manifest = {
        "schema_version": CASE_SCHEMA_VERSION,
        "ai_generated_content_notice": AI_GENERATED_CONTENT_NOTICE,
        "template_version": TEMPLATE_VERSION,
        "case_id": "SYNTH_DEMO_T0",
        "company_name": "合成演示企业（仅用于系统验收）",
        "company_alias": "合成案例",
        "ticker": "",
        "t0": "2025-04-30",
        "currency": "CNY",
        "amount_unit": "元",
        "statement_scope": "合并",
        "sample_type": "synthetic",
        "model_transfer_allowed": False,
        "retention_expires_at": "2026-12-31",
        "documents": documents,
    }
    financial = io.StringIO()
    writer = csv.writer(financial, lineterminator="\n")
    writer.writerow(
        [
            "case_id",
            "evidence_id",
            "field_kind",
            "year",
            "value",
            "unit",
            "statement_scope",
            "field_basis",
            "document_id",
            "pdf_page",
            "print_page",
            "locator",
        ]
    )
    synthetic_values = {
        2024: {"revenue": 125000000, "accounts_receivable": 41000000, "operating_cash_flow": 8500000, "net_profit": 9300000},
        2023: {"revenue": 110000000, "accounts_receivable": 30000000, "operating_cash_flow": 7800000, "net_profit": 8100000},
        2022: {"revenue": 100000000, "accounts_receivable": 25000000, "operating_cash_flow": 7200000, "net_profit": 7500000},
    }
    for year, fields in synthetic_values.items():
        for kind, value in fields.items():
            writer.writerow(
                [
                    "SYNTH_DEMO_T0",
                    f"SYNTH_DEMO_T0_{FIELD_PREFIX[kind]}_{year}",
                    kind,
                    year,
                    value,
                    "元",
                    "合并",
                    "net" if kind == "accounts_receivable" else "reported",
                    f"SYNTH-AR-{year}",
                    1,
                    1,
                    "合成模板 / 字段登记表（不是 PDF 自动取数）",
                ]
            )
    readme = """审迹智链标准案例包模板\n\n本包为合成验收样例，不是第二个正式案例，也不包含实验结论。\n导入器只读取 financial_fields 中已登记并可回查的字段；不会从任意 PDF 自动取数。\n正式使用前应更换为合法公开或已授权脱敏资料、更新哈希，并由会计专业人员复核口径。\n"""
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("case_manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
        archive.writestr("financial_fields.csv", financial.getvalue().encode("utf-8-sig"))
        archive.writestr("aging.csv", "customer_id,aging_bucket,amount,as_of_date,evidence_note\n")
        archive.writestr("subsequent_receipts.csv", "customer_id,receipt_date,amount,reference,evidence_note\n")
        archive.writestr("contract_terms.csv", "contract_id,customer_id,credit_days,settlement_terms,acceptance_terms,evidence_note\n")
        archive.writestr("deidentification_log.csv", "field_or_file,action,reviewer,status,note\n")
        archive.writestr("README_填写说明.txt", readme.encode("utf-8-sig"))
        for year, content in pdfs.items():
            archive.writestr(f"documents/synthetic_annual_report_{year}.pdf", content)
    return output.getvalue()
