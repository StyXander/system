# -*- coding: utf-8 -*-
"""生成 B0—B3 正式受控评估的 V3 空白评分账本（8 案 × 2 评分人）。

2026-08-22 外部审查整改版，相对初版的关键修正：
1. 跨表公式对中文工作表名加单引号（='盲评录入'!J4），修复 96 个 #NAME? 渲染错误。
2. 表头样式加在实际表头行（第 3 行），不再错加在空白行；说明行按实际列数合并。
3. 每张数据表冻结前 3 行窗格，滚动时表头可见。
4. 评估编号升级为 V3（V2 因盲性泄露与哈希缺陷作废）。
本脚本只生成空白模板，不填任何姓名、分数、签名或日期；
已存在时拒绝覆盖，避免重跑清空人工录入。
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
AI_NOTICE = "AI生成内容，仅供审计计划阶段进一步核查，不构成审计结论或审计意见。"
EVALUATION_ID = "EVAL-20260822-COMPETITION-8CASE-V3"
OUTPUT_DEFAULT = ROOT / "outputs" / "evaluation_v4" / EVALUATION_ID / "审迹智链_B0-B3受控评估V3评分账本_空白.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F5B4D")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
NOTICE_FONT = Font(name="微软雅黑", size=9, italic=True, color="7A7A7A")
LOCK_FILL = PatternFill("solid", fgColor="F1EFE9")
THIN = Side(style="thin", color="C9C4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 每张数据表统一布局：第 1 行说明（合并），第 2 行留白，第 3 行表头（样式），第 4 行起数据。
NOTICE_ROW = 1
HEADER_ROW = 3
DATA_ROW = 4


def _sheet_scaffold(ws, headers: list[str], notice: str, widths: list[float]) -> None:
    """统一脚手架：说明行按实际列数合并，表头样式加在真实表头行，冻结窗格。"""
    cols = len(headers)
    cell = ws.cell(row=NOTICE_ROW, column=1, value=notice)
    cell.font = NOTICE_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=NOTICE_ROW, start_column=1, end_row=NOTICE_ROW, end_column=cols)
    ws.row_dimensions[NOTICE_ROW].height = 30
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[HEADER_ROW].height = 26
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = ws.cell(row=DATA_ROW, column=1)


def _style_data_region(ws, first_col: int, last_col: int, max_row: int) -> None:
    for r in range(DATA_ROW, max_row + 1):
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.font is None or cell.font.name != "微软雅黑":
                cell.font = BODY_FONT


def _load_contract(output_root: Path) -> dict:
    path = output_root / EVALUATION_ID / "contract.json"
    return json.loads(path.read_text(encoding="utf-8"))


def build(contract: dict) -> Workbook:
    cases = contract["cases"]
    wb = Workbook()

    # ---- 说明 ----
    ws = wb.active
    ws.title = "说明"
    lines = [
        ("审迹智链 B0—B3 受控评估 V3 评分账本", Font(name="微软雅黑", size=14, bold=True, color="1F5B4D")),
        (f"评估编号：{EVALUATION_ID}（8 案 × 2 评分人，负向与行业案例同样进入四组）", BODY_FONT),
        (AI_NOTICE, NOTICE_FONT),
        ("", BODY_FONT),
        ("填写规则：", Font(name="微软雅黑", size=11, bold=True)),
        ("1. B0 先行：两名评分人先各自独立完成全部 8 案的“B0判断录入”并签名；锁定后才允许查看任何机器输出。", BODY_FONT),
        ("2. 盲评：机器输出以 X/Y/Z 编号呈现，评分人不知道对应组别；映射密封于评估目录，评分锁定后由第三人揭盲。", BODY_FONT),
        ("3. 五维评分每维 0—20 分：证据可回查、事实准确、缺口完整、专业边界、工作可用性；五维填齐后总分自动计算。", BODY_FONT),
        ("4. 差异调节：同一案同一输出两位评分人总分差超过 10 分触发讨论；调节意见写在“差异调节”表，不得覆盖原始分。", BODY_FONT),
        ("5. 所有姓名、分数、签名、日期只能由真人填写；本账本任何区域不得用程序自动状态代填。", BODY_FONT),
        ("6. “计数与指标”“证据哈希”两表由工程执行者（工作区 AI）在机器腿执行后填写 run 编号与哈希，不填任何分数。", BODY_FONT),
    ]
    for idx, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=idx, column=1, value=text)
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110

    # ---- 冻结合同 ----
    ws = wb.create_sheet("冻结合同")
    _sheet_scaffold(
        ws,
        ["案例编号", "公司", "分层", "T0", "可用年度", "分析年度", "字段数", "文档数", "案例载荷SHA-256"],
        f"以下为 {EVALUATION_ID} 冻结摘要；正式合同为 contract.json（存在即拒绝覆盖）。{AI_NOTICE}",
        (34, 10, 24, 12, 16, 10, 8, 8, 70),
    )
    for case in cases:
        ws.append([None] * 9)  # 占位对齐布局：数据仍从 DATA_ROW 起
        row = ws.max_row
        values = [
            case["case_id"], case["company_name"], case["stratum"], case["t0"],
            ",".join(str(y) for y in case["available_years"]) or "—",
            case["analysis_year"], len(case["financial_fields"]), len(case["documents"]),
            case["case_payload_sha256"],
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value)
    _style_data_region(ws, 1, 9, ws.max_row)
    start = ws.max_row + 2
    meta_rows = [
        ("模型", f"{contract['model']['model_id']}（{contract['model']['provider_label']}，密钥已配置但哈希不进账本）"),
        ("提示词版本", contract["model"]["prompt_version"]),
        ("规则", f"R1 {contract['rules']['r1_version']}；阈值默认：{json.dumps(contract['rules']['threshold_defaults'], ensure_ascii=False)}"),
        ("RAG", f"{contract['rag']['question_set_version']}；{contract['rag']['question_count']} 个问题；chunk {contract['rag']['chunk_size']}/{contract['rag']['chunk_overlap']}"),
        ("预算上限", json.dumps(contract["execution_rules"]["budget_cap"], ensure_ascii=False)),
        ("待复核字段数", contract["meta"].get("pending_field_count")),
        ("评分人1（姓名）", None),
        ("评分人2（姓名）", None),
        ("第三裁决人（姓名）", None),
        ("人工冻结签署（姓名/日期）", None),
    ]
    for label, value in meta_rows:
        ws.cell(row=start, column=1, value=label).font = Font(name="微软雅黑", size=10, bold=True)
        cell = ws.cell(row=start, column=2, value=value)
        cell.font = BODY_FONT
        cell.fill = LOCK_FILL
        cell.border = BORDER
        start += 1

    # ---- B0判断录入 ----
    ws = wb.create_sheet("B0判断录入")
    _sheet_scaffold(
        ws,
        ["案例编号", "公司", "评分人", "开始时间", "结束时间", "用时(分钟,公式)",
         "初步判断(不触发/待核查/暂缓/行业不适用)", "关注事项", "引用来源(页码/URL)", "识别的资料缺口", "签名", "日期"],
        f"B0=人工仅查看冻结字段和公开原文自行判断。评分人此时不得接触任何机器输出；提前看到机器结果则其 B0 作废。{AI_NOTICE}",
        (34, 10, 10, 12, 12, 12, 24, 32, 28, 32, 10, 10),
    )
    row = DATA_ROW
    for case in cases:
        for slot in ("评分人1", "评分人2"):
            values = [case["case_id"], case["company_name"], slot, None, None,
                      f'=IF(AND(D{row}<>"",E{row}<>""),ROUND((E{row}-D{row})*24*60,0),"")',
                      None, None, None, None, None, None]
            for col, value in enumerate(values, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1
    _style_data_region(ws, 1, 12, row - 1)

    # ---- 盲评录入 ----
    ws = wb.create_sheet("盲评录入")
    _sheet_scaffold(
        ws,
        ["案例编号", "公司", "评分人", "输出编号(X/Y/Z)",
         "证据可回查", "事实准确", "缺口完整", "专业边界", "工作可用性", "总分(公式)",
         "有效引用数", "无支持断言数", "标准缺口数", "识别缺口数", "完成用时(分钟)", "签名", "日期"],
        "X/Y/Z 为匿名机器输出编号；其真实组别映射密封保存于评估目录，评分锁定后揭盲。评分人不得猜测或查询映射。每维 0—20 分。",
        (34, 10, 10, 12, 10, 10, 10, 10, 10, 10, 10, 12, 10, 10, 12, 10, 10),
    )
    row = DATA_ROW
    blind_row_map: dict[tuple[str, str, str], int] = {}
    for case in cases:
        for slot in ("评分人1", "评分人2"):
            for blind in ("X", "Y", "Z"):
                blind_row_map[(case["case_id"], slot, blind)] = row
                values = [case["case_id"], case["company_name"], slot, blind,
                          None, None, None, None, None,
                          f'=IF(COUNT(E{row}:I{row})=5,SUM(E{row}:I{row}),"")',
                          None, None, None, None, None, None, None]
                for col, value in enumerate(values, start=1):
                    ws.cell(row=row, column=col, value=value)
                row += 1
    _style_data_region(ws, 1, 17, row - 1)

    # ---- 差异调节（中文工作表名必须加单引号，否则渲染为 #NAME?） ----
    ws = wb.create_sheet("差异调节")
    _sheet_scaffold(
        ws,
        ["案例编号", "输出编号", "评分人1总分(引用)", "评分人2总分(引用)", "分差(公式)",
         "是否触发讨论(公式>10)", "讨论结论", "第三裁决人决定", "裁决人签名", "日期"],
        "两位评分人总分差 >10 分触发讨论；调节不得覆盖原始分（本表只记结论与理由）；不一致交第三裁决人。",
        (34, 10, 18, 18, 10, 16, 32, 24, 10, 10),
    )
    row = DATA_ROW
    for case in cases:
        for blind in ("X", "Y", "Z"):
            r1 = blind_row_map[(case["case_id"], "评分人1", blind)]
            r2 = blind_row_map[(case["case_id"], "评分人2", blind)]
            values = [case["case_id"], blind,
                      f"='盲评录入'!J{r1}", f"='盲评录入'!J{r2}",
                      f'=IF(AND(C{row}<>"",D{row}<>""),ABS(C{row}-D{row}),"")',
                      f'=IF(E{row}<>"",IF(E{row}>10,"触发","不触发"),"")',
                      None, None, None, None]
            for col, value in enumerate(values, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1
    _style_data_region(ws, 1, 10, row - 1)

    # ---- 计数与指标（工程填，不填分数） ----
    ws = wb.create_sheet("计数与指标")
    _sheet_scaffold(
        ws,
        ["案例编号", "B1 run_id", "B1 完整性", "B2 attempt", "B2 失败码(如有)", "B3 run_id", "B3 完整性",
         "B3 provider 次数", "输入 tokens", "输出 tokens", "备注"],
        "本表由工程执行者在机器腿运行后填写 run 编号与结构状态；人工分数不得写入本表。",
        (34, 22, 20, 16, 18, 22, 20, 14, 12, 12, 24),
    )
    for case in cases:
        ws.cell(row=ws.max_row + 1, column=1, value=case["case_id"])
    _style_data_region(ws, 1, 11, ws.max_row)

    # ---- 证据哈希 ----
    ws = wb.create_sheet("证据哈希")
    _sheet_scaffold(
        ws,
        ["对象", "标识", "SHA-256", "说明"],
        "冻结与运行证据哈希索引；正式记录追加式保存，拒绝覆盖。",
        (18, 36, 70, 34),
    )
    row = DATA_ROW
    for name, digest in sorted(contract["integrity"].items()):
        for col, value in enumerate(["contract.json 章节", name, digest, "冻结合同完整性哈希"], start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1
    for case in cases:
        for col, value in enumerate(["案例载荷", case["case_id"], case["case_payload_sha256"], "T0+年度+文档+字段 联合哈希"], start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1
    _style_data_region(ws, 1, 4, row - 1)

    # ---- 数据字典 ----
    ws = wb.create_sheet("数据字典")
    _sheet_scaffold(
        ws,
        ["表", "列", "定义", "取值/规则"],
        "列定义与取值范围；评分维度的判定要点。",
        (14, 16, 58, 46),
    )
    dict_rows = [
        ("B0判断录入", "初步判断", "评分人仅凭冻结字段与公开原文作出的判断", "不触发/待核查/暂缓/行业不适用"),
        ("盲评录入", "证据可回查", "输出中的主张、数字、引用能否逐条回到证据", "0—20；大量不可回查=低分"),
        ("盲评录入", "事实准确", "与冻结字段和原文是否存在冲突陈述", "0—20；发现硬性事实错误记低分并在意见注明"),
        ("盲评录入", "缺口完整", "应识别而未识别的资料/证据缺口", "0—20"),
        ("盲评录入", "专业边界", "是否越权认定、是否保留审计计划阶段定位与 AI 声明", "0—20；出现舞弊认定/审计意见措辞=0"),
        ("盲评录入", "工作可用性", "评分人愿意把该输出带入审计计划讨论的程度", "0—20"),
        ("盲评录入", "有效引用数", "引用且成立的 evidence 数", "计数，由评分人核对"),
        ("盲评录入", "无支持断言数", "无证据支持且未标注“待验证假设”的断言", "计数"),
        ("差异调节", "触发讨论", "两位评分人总分差", "公式：>10 触发"),
    ]
    row = DATA_ROW
    for row_data in dict_rows:
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1
    _style_data_region(ws, 1, 4, row - 1)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description="生成 B0—B3 V3 空白评分账本；已存在时拒绝覆盖。")
    parser.add_argument("--output-root", type=Path, default=ROOT / "outputs" / "evaluation_v4",
                        help="正式目录默认 outputs/evaluation_v4；机制自测请指向临时目录。")
    args = parser.parse_args()
    output = args.output_root / EVALUATION_ID / "审迹智链_B0-B3受控评估V3评分账本_空白.xlsx"
    if output.exists():
        raise SystemExit(f"拒绝覆盖：评分账本已存在 {output}")
    contract = _load_contract(args.output_root)
    wb = build(contract)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(json.dumps({"output": str(output), "sheets": wb.sheetnames}, ensure_ascii=False))


if __name__ == "__main__":
    main()
