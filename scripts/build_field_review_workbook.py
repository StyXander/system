# -*- coding: utf-8 -*-
"""生成 8 案字段回页复核工作底稿（人工复核入口，先于正式 V3 冻结）。

背景：V2 合同把自动提取候选值直接冻结，其中至少三处已被外部审查对照官方
年报确认错误（万科A/中国中铁/中国铁建的应收账款）。按变更规则不能修改
旧合同冒充原冻结，必须先由真人逐字段回页复核，再生成 V3 正式合同。

本底稿列出全部字段的当前候选值、异常标记与来源页，并附外部审查参考值
（仍属 AI 核对线索，须回页确认）。回页值/决策/复核人/日期列只能真人填写；
生成器不填任何复核结论。已存在时拒绝覆盖。
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[1]
AI_NOTICE = "AI生成内容，仅供审计计划阶段进一步核查，不构成审计结论或审计意见。"
V2_CONTRACT = ROOT / "outputs" / "evaluation_v4" / "EVAL-20260822-COMPETITION-8CASE-V2" / "contract.json"
OUTPUT_DEFAULT = ROOT / "outputs" / "evaluation_v4" / "field-review" / "字段回页复核工作底稿_2026-08-22.xlsx"

# 外部审查对照官方年报给出的参考值（AI 核对线索，仍须回页确认）。
EXTERNAL_REFERENCES = {
    ("CNINFO_000002_T0_20260331", "accounts_receivable", 2024): "8,110,758,258.05（官方2025年报比较数，外部审查核对）",
    ("CNINFO_000002_T0_20260331", "accounts_receivable", 2023): "官方2025年报比较数显示为数十亿元量级；当前候选 4.00 元明显异常",
    ("CNINFO_601390_T0_20260330", "accounts_receivable", 2023): "当前候选 -5,000 元为负数，不能确认",
    ("CNINFO_601186_T0_20260330", "accounts_receivable", 2024): "204,759,238 千元（官方2025年报比较数，外部审查核对）",
}

FIELD_KIND_LABELS = {
    "revenue": "营业收入",
    "accounts_receivable": "应收账款",
    "operating_cash_flow": "经营活动现金流量净额",
    "net_profit": "净利润",
}

HEADER_FILL = PatternFill("solid", fgColor="1F5B4D")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
NOTICE_FONT = Font(name="微软雅黑", size=9, italic=True, color="7A7A7A")
LOCK_FILL = PatternFill("solid", fgColor="FDF6E3")
THIN = Side(style="thin", color="C9C4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _magnitude_flag(value, unit: str) -> str:
    if not isinstance(value, (int, float)):
        return ""
    if value < 0:
        return "负值：应收/收入类字段出现负数，须回页确认"
    if abs(value) < 10000 and unit == "元":
        return "数量级异常：大型上市公司金额低于 1 万元，疑似行列取数错误"
    return ""


def main() -> None:
    parser = argparse.ArgumentParser(description="生成字段回页复核工作底稿；已存在时拒绝覆盖。")
    parser.add_argument("--output", type=Path, default=OUTPUT_DEFAULT)
    args = parser.parse_args()
    if args.output.exists():
        raise SystemExit(f"拒绝覆盖：工作底稿已存在 {args.output}")

    contract = json.loads(V2_CONTRACT.read_text(encoding="utf-8"))
    wb = Workbook()

    # ---- 修正落地路径说明（复核人必读） ----
    guide = wb.create_sheet("修正落地路径说明")
    wb.remove(wb["Sheet"])  # 移除 Workbook 默认空表，交付底稿不留空白页
    guide_lines = [
        ("字段复核后，修正值如何生效（复核人必读）", Font(name="微软雅黑", size=13, bold=True, color="1F5B4D")),
        ("", BODY_FONT),
        ("1. CNINFO 案例（万科A、中国中铁、中国铁建、五粮液、中国海油、中国建筑、中国人寿）：", Font(name="微软雅黑", size=11, bold=True)),
        ("   由工程侧调用字段确认接口逐条录入复核结论：POST /api/cases/{case_id}/fields/confirm，", BODY_FONT),
        ("   decision 三选一：confirm（与原件一致）/ correct（按回页值修正，附修正前值与原因）/ reject（原件不支持）。", BODY_FONT),
        ("   复核人姓名、日期、原因随接口登记并进入历史记录；工作底稿由工程侧照此转录，不代填。", BODY_FONT),
        ("2. 标准股份（STD_DEV_T0）：字段冻结在 backend/app/data.py 注册表内。", Font(name="微软雅黑", size=11, bold=True)),
        ("   按注册表自身规则，修改冻结金额必须附带新的来源复核记录和快照版本，不允许原地悄改；", BODY_FONT),
        ("   因此 STD 字段的修正属于输入变化，由工程侧按复核结论修改 data.py 并升快照版本，", BODY_FONT),
        ("   随后以新评估编号冻结（不修改既有合同冒充原冻结）。", BODY_FONT),
        ("3. 全部字段复核完成后：工程侧重新运行 freeze_evaluation_v4.py 生成正式 V3 合同", Font(name="微软雅黑", size=11, bold=True)),
        ("   （此时不得再出现待复核字段），再生成 B0 盲包与评分账本。", BODY_FONT),
        ("4. 本底稿的“回页核对值/决策/复核人/日期”四列是复核过程的原始留痕，转录后随评估目录封存。", BODY_FONT),
        ("", BODY_FONT),
        (f"统一 AI 声明：{AI_NOTICE}", NOTICE_FONT),
    ]
    for idx, (text, font) in enumerate(guide_lines, start=1):
        cell = guide.cell(row=idx, column=1, value=text)
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    guide.column_dimensions["A"].width = 110

    ws = wb.create_sheet("字段回页复核")

    headers = [
        "案例编号", "公司", "字段类别", "年度", "当前候选值", "单位", "冻结状态",
        "异常标记", "外部审查参考值(仍须回页)", "来源 URL", "PDF 页", "定位",
        "回页核对值(真人填)", "决策(确认/修正/拒绝)", "复核人(真人填)", "日期(真人填)",
    ]
    widths = [34, 10, 12, 8, 18, 8, 22, 30, 40, 44, 8, 24, 18, 16, 12, 12]
    notice = (
        f"用途：真人逐字段回页复核后，才能生成正式 V3 冻结合同。{AI_NOTICE}\n"
        "“回页核对值/决策/复核人/日期”四列只能由真人填写；外部审查参考值是 AI 核对线索，不构成确认。"
    )
    cell = ws.cell(row=1, column=1, value=notice)
    cell.font = NOTICE_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 34
    for col_idx, (title, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[3].height = 28
    ws.freeze_panes = "A4"

    row = 4
    total = 0
    for case in contract["cases"]:
        for f in sorted(case["financial_fields"], key=lambda x: (x.get("year") or 0, str(x.get("field_kind"))), reverse=True):
            doc = next((d for d in case["documents"] if d.get("document_id") == f.get("document_id")), {})
            reference = EXTERNAL_REFERENCES.get((case["case_id"], f.get("field_kind"), f.get("year")), "")
            values = [
                case["case_id"], case["company_name"],
                FIELD_KIND_LABELS.get(f.get("field_kind"), str(f.get("field_kind"))),
                f.get("year"),
                f"{f.get('value'):,.2f}" if isinstance(f.get("value"), (int, float)) else str(f.get("value")),
                f.get("unit") or "元",
                f.get("field_freeze_status") if "field_freeze_status" in f else "candidate_pending_human_review",
                _magnitude_flag(f.get("value"), f.get("unit") or "元"),
                reference,
                doc.get("source_url") or "—",
                f.get("pdf_page") or "—",
                f.get("locator") or "—",
                None, None, None, None,
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = BORDER
                cell.font = BODY_FONT
                if col >= 13:
                    cell.fill = LOCK_FILL
            row += 1
            total += 1

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(json.dumps({"output": str(args.output), "fields": total}, ensure_ascii=False))


if __name__ == "__main__":
    main()
